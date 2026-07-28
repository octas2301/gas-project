#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M2: 案L CSV または GENERATED offer → SC公式 ListingLoader xlsm 埋め

- テンプレは人間が SC から DL した公式ファイルを指定（自動DLしない）
- SC への自動アップロードはしない
- C1 HPC ツールは使わない
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

LOG = logging.getLogger("m2_loader_fill")
SCRIPT_DIR = Path(__file__).resolve().parent

try:
    from openpyxl import load_workbook
except ImportError as e:  # pragma: no cover
    raise SystemExit("openpyxl が必要です: pip install openpyxl") from e


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def resolve_path(p: str, base: Path) -> Path:
    path = Path(p)
    if not path.is_absolute():
        path = (base / path).resolve()
    return path


def _load_json(path: Path) -> Dict[str, Any]:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _save_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")


def load_offer_csv(path: Path) -> List[Dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("offer CSV が空です: %s" % path)
        rows = []
        for raw in reader:
            row = {k: str(raw.get(k) or "").strip() for k in reader.fieldnames}
            if row.get("sku"):
                rows.append(row)
        return rows


def load_generated_offers(path: Path) -> List[Dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("GENERATED が空です: %s" % path)
        rows = []
        for raw in reader:
            row = {k: str(raw.get(k) or "").strip().replace("\r", "") for k in reader.fieldnames}
            if (row.get("track") or "").upper() == "A" and (row.get("variationRole") or "").lower() == "offer":
                rows.append(row)
        return rows


def build_key_to_col(ws, header_key_row: int, max_col: int = 120) -> Dict[str, int]:
    """行 header_key_row のフィールド名 → 1-based 列番号。"""
    out: Dict[str, int] = {}
    for col in range(1, max_col + 1):
        v = ws.cell(header_key_row, col).value
        if v is None:
            continue
        key = str(v).strip()
        if key and key not in out:
            out[key] = col
    return out


def map_row_values(
    src: Dict[str, str],
    field_map: Dict[str, str],
    defaults: Dict[str, str],
) -> Dict[str, str]:
    """ListingLoader フィールド名 → 埋め値。"""
    values = dict(defaults)
    for field_key, src_col in field_map.items():
        val = src.get(src_col) or ""
        if val != "":
            values[field_key] = val
    return values


def validate_offer_like(sku: str, asin: str, price: str, line_no: int) -> Optional[str]:
    if not sku:
        return "line %d: sku 空" % line_no
    if not re.match(r"^B0[A-Z0-9]{8}$", asin, re.I):
        return "line %d: ASIN不正 %r sku=%s" % (line_no, asin, sku)
    try:
        if float(price) <= 0:
            return "line %d: price 不正 %r" % (line_no, price)
    except ValueError:
        return "line %d: price 非数値 %r" % (line_no, price)
    return None


def fill_workbook(
    template_path: Path,
    out_path: Path,
    rows_values: List[Dict[str, str]],
    sheet_name: str,
    header_key_row: int,
    data_start_row: int,
    dry_run: bool,
) -> Dict[str, Any]:
    wb = load_workbook(template_path, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError("シートがありません: %s / sheets=%s" % (sheet_name, wb.sheetnames))
    ws = wb[sheet_name]
    key_to_col = build_key_to_col(ws, header_key_row)
    missing_keys = []
    for rv in rows_values:
        for k in rv:
            if k not in key_to_col:
                missing_keys.append(k)
    missing_keys = sorted(set(missing_keys))
    if missing_keys:
        LOG.warning("テンプレに無いフィールド（スキップ）: %s", missing_keys)

    applied: List[Dict[str, Any]] = []
    for i, rv in enumerate(rows_values):
        row_idx = data_start_row + i
        cell_writes = {}
        for field_key, val in rv.items():
            col = key_to_col.get(field_key)
            if not col:
                continue
            if not dry_run:
                ws.cell(row_idx, col).value = val
            cell_writes[field_key] = {"col": col, "value": val}
        applied.append({"excelRow": row_idx, "writes": cell_writes})

    meta = {
        "template": str(template_path),
        "output": str(out_path),
        "sheet": sheet_name,
        "matchedKeys": sorted(key_to_col.keys())[:5],
        "keyCount": len(key_to_col),
        "missingKeys": missing_keys,
        "rows": applied,
    }

    if dry_run:
        wb.close()
        return meta

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.resolve() != template_path.resolve():
        # keep_vba 保存はコピー先へ
        pass
    wb.save(out_path)
    wb.close()
    return meta


def run(
    template: Path,
    output: Path,
    map_path: Path,
    offer_csv: Optional[Path],
    generated_csv: Optional[Path],
    mode: str,
) -> int:
    colmap = _load_json(map_path)
    sheet = str(colmap.get("template_sheet") or "テンプレート")
    header_key_row = int(colmap.get("header_key_row") or 5)
    data_start_row = int(colmap.get("data_start_row") or 7)
    defaults = dict(colmap.get("defaults") or {})

    source_label = ""
    raw_rows: List[Dict[str, str]] = []
    field_map: Dict[str, str] = {}

    if offer_csv:
        source_label = str(offer_csv)
        raw_rows = load_offer_csv(offer_csv)
        field_map = dict(colmap.get("offer_csv_to_field") or {})
        get_sku = lambda r: r.get("sku") or ""
        get_asin = lambda r: r.get("product-id") or ""
        get_price = lambda r: r.get("price") or ""
    elif generated_csv:
        source_label = str(generated_csv)
        raw_rows = load_generated_offers(generated_csv)
        field_map = dict(colmap.get("generated_to_field") or {})
        get_sku = lambda r: r.get("sellerSku") or ""
        get_asin = lambda r: r.get("asin") or ""
        get_price = lambda r: r.get("priceAmazon") or ""
    else:
        raise SystemExit("--offer-csv または --generated のどちらかが必須です")

    if not raw_rows:
        LOG.error("埋め対象行が0件です: %s", source_label)
        return 2

    rows_values: List[Dict[str, str]] = []
    errors: List[str] = []
    for i, src in enumerate(raw_rows, start=1):
        err = validate_offer_like(get_sku(src), get_asin(src), get_price(src), i)
        if err:
            errors.append(err)
            continue
        rows_values.append(map_row_values(src, field_map, defaults))

    report = {
        "runId": "M2_LOADER_FILL_%s" % _utc_stamp(),
        "mode": mode,
        "version": "M2-LoaderFill-v1",
        "source": source_label,
        "template": str(template),
        "counts": {
            "in": len(raw_rows),
            "accepted": len(rows_values),
            "rejected": len(errors),
        },
        "errors": errors,
    }

    if not rows_values:
        LOG.error("受理0件。errors=%s", errors)
        return 2

    dry_run = mode == "dry_run"
    if dry_run:
        # dry_run でも出力パスはレポート用に仮置き
        out_path = output
    else:
        out_path = output
        if out_path.exists() and out_path.resolve() == template.resolve():
            raise SystemExit("output が template と同じです。別名を指定してください")
        if not out_path.suffix:
            out_path = out_path.with_suffix(".xlsm")
        # テンプレをコピーしてから開くと VBA 保全が安定
        if out_path.resolve() != template.resolve():
            shutil.copy2(template, out_path)
            template_for_open = out_path
        else:
            template_for_open = template
        meta = fill_workbook(
            template_for_open,
            out_path,
            rows_values,
            sheet,
            header_key_row,
            data_start_row,
            dry_run=False,
        )
        report["fill"] = meta
        report["outputPath"] = str(out_path)
        LOG.info("出力: %s rows=%d", out_path, len(rows_values))
        _write_report(out_path.parent, report)
        return 0 if not errors else 0

    # dry_run: テンプレを読みキー照合のみ
    meta = fill_workbook(
        template,
        output,
        rows_values,
        sheet,
        header_key_row,
        data_start_row,
        dry_run=True,
    )
    report["fill"] = meta
    report["outputPath"] = "(dry_run)"
    LOG.info("dry_run OK accepted=%d missingKeys=%s", len(rows_values), meta.get("missingKeys"))
    for row in meta.get("rows") or []:
        w = row.get("writes") or {}
        sku = (w.get("contribution_sku#1.value") or {}).get("value")
        asin = (w.get("merchant_suggested_asin#1.value") or {}).get("value")
        price = None
        for k, cell in w.items():
            if "our_price" in k:
                price = cell.get("value")
                break
        LOG.info("  excelRow=%s sku=%s asin=%s price=%s", row.get("excelRow"), sku, asin, price)
    _write_report(output.parent if output.suffix else Path("."), report)
    return 0 if not errors else 0


def _write_report(dir_path: Path, report: Dict[str, Any]) -> None:
    try:
        dir_path.mkdir(parents=True, exist_ok=True)
        p = dir_path / ("%s_REPORT.json" % report["runId"])
        _save_json(p, report)
        LOG.info("レポート: %s", p)
    except OSError as e:
        LOG.warning("レポート保存失敗: %s", e)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="M2: 公式 ListingLoader xlsm に offer 行を埋める（SC自動UPなし）"
    )
    parser.add_argument("--template", required=True, help="SCからDLした ListingLoader .xlsm")
    parser.add_argument("--output", required=True, help="埋め後の出力 .xlsm パス")
    parser.add_argument("--offer-csv", default=None, help="*_M2_OFFER_LOADER.csv")
    parser.add_argument("--generated", default=None, help="*_GENERATED.csv（track=A offer）")
    parser.add_argument(
        "--map",
        default=str(SCRIPT_DIR / "listing_loader_map.json"),
        help="列マップ JSON",
    )
    parser.add_argument("--mode", choices=["dry_run", "prod"], default="dry_run")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    template = Path(args.template).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()
    map_path = Path(args.map).expanduser().resolve()
    offer = Path(args.offer_csv).expanduser().resolve() if args.offer_csv else None
    generated = Path(args.generated).expanduser().resolve() if args.generated else None

    if not template.is_file():
        raise SystemExit("template がありません: %s" % template)
    if not map_path.is_file():
        raise SystemExit("map がありません: %s" % map_path)
    if offer and not offer.is_file():
        raise SystemExit("offer-csv がありません: %s" % offer)
    if generated and not generated.is_file():
        raise SystemExit("generated がありません: %s" % generated)

    return run(template, output, map_path, offer, generated, args.mode)


if __name__ == "__main__":
    sys.exit(main())
