# -*- coding: utf-8 -*-
"""シート数量と提出xlsxの結合検証。"""
from __future__ import annotations

import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from paths import folder_path, latest_xlsx, load_config
from qty_logic import compute_q_deal, deal_day_count
from sheet_schema import LANE_B, MASTER_SHEET, SALE_SHEET
from sheets_io import read_sheet_rows, sheets_service
from template_parse import detect_header, find_template_sheet, load_column_map

try:
    import openpyxl
except ImportError:
    print("openpyxl required")
    raise SystemExit(2)


def main() -> int:
    cfg = load_config(HERE / "config.local.json")
    svc = sheets_service(write=False)
    sid = str(cfg["ads_spreadsheet_id"])
    fails = []

    mh, master = read_sheet_rows(svc, sid, MASTER_SHEET)
    sh, sales = read_sheet_rows(svc, sid, SALE_SHEET)
    print("master headers has V30", "V30" in mh, "sale has V30", "V30" in sh)

    b_rows = [r for r in sales if str(r.get("レーン") or "") == LANE_B]
    print(f"B rows={len(b_rows)}")

    for r in b_rows:
        asin = str(r.get("ASIN") or "").upper()
        try:
            v30 = float(str(r.get("V30")).replace(",", "")) if str(r.get("V30") or "").strip() else None
        except ValueError:
            v30 = None
        # Q_fba from master
        q_fba = None
        for m in master:
            if str(m.get("ASIN") or "").upper() == asin:
                try:
                    q_fba = float(str(m.get("Q_fba")).replace(",", "")) if str(m.get("Q_fba") or "").strip() else None
                except ValueError:
                    q_fba = None
                break
        d = deal_day_count(r.get("開始日"), r.get("終了日"))
        expect = compute_q_deal(
            v30=v30,
            d_days=d,
            schedule=str(r.get("スケジュール") or ""),
            q_fba=q_fba,
        )
        got = int(float(str(r.get("販売商品数_確定") or "0").replace(",", "") or "0"))
        ok = got == expect["Q_deal"]
        line = (
            f"{asin} {r.get('開始日')} V30={v30} D={d} expect={expect['Q_deal']} got={got} "
            f"SC={r.get('販売商品数_SC')}"
        )
        print(("OK  " if ok else "FAIL") + ":", line)
        if not ok:
            fails.append(line)

    # 提出xlsx
    folder03 = folder_path(cfg, "03")
    submits = sorted(
        [p for p in folder03.glob("DEALS_*_submit.xlsx") if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not submits:
        fails.append("提出xlsxなし")
        print("FAIL: 提出xlsxなし")
        return 1
    src = submits[0]
    print("submit:", src.name)
    audit_path = src.with_name(src.name.replace("_submit.xlsx", "_audit.json"))
    if audit_path.is_file():
        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        summary = audit.get("summary") or {}
        print("audit summary:", summary)
        if summary.get("deal_type_policy") and "おすすめ" not in str(summary.get("deal_type_policy")):
            fails.append("deal_type_policy")

    cmap = load_column_map()
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = find_template_sheet(wb, list(cmap.get("template_sheet_name_candidates") or []))
    _hr, data_start, colmap = detect_header(ws, cmap)
    opt_in = []
    for r in range(data_start, (ws.max_row or data_start) + 1):
        asin = str(ws.cell(r, colmap["deal_asin"]).value or "").strip().upper()
        if not asin:
            continue
        part = str(ws.cell(r, colmap["participating"]).value or "").strip()
        dtype = str(ws.cell(r, colmap.get("deal_type", 1)).value or "").strip()
        if part in ("はい", "YES", "Yes", "true", "TRUE"):
            qty = ws.cell(r, colmap["committed_units"]).value if colmap.get("committed_units") else None
            opt_in.append({"row": r, "asin": asin, "deal_type": dtype, "qty": qty, "sku": ws.cell(r, colmap["sku"]).value})
            if "数量限定" in dtype:
                fails.append(f"数量限定が参加中 row={r} asin={asin}")
                print("FAIL: 数量限定 opt_in", asin, dtype)
            if "おすすめ" not in dtype:
                fails.append(f"非おすすめ opt_in row={r} {dtype}")
                print("FAIL: 非おすすめ", dtype)

    print(f"opt_in rows={len(opt_in)}")
    for o in opt_in:
        print(" ", o)

    # シートBのSmile枠数量とxlsx数量の突合（ASIN単位・直近Smile）
    smile_sheet = {
        str(r.get("ASIN")).upper(): int(float(str(r.get("販売商品数_確定"))))
        for r in b_rows
        if "Smile" in str(r.get("スケジュール") or "") or "スマイル" in str(r.get("スケジュール") or "")
    }
    for o in opt_in:
        asin = o["asin"]
        if asin in smile_sheet and o["qty"] is not None:
            try:
                q = int(float(o["qty"]))
            except (TypeError, ValueError):
                continue
            if q != smile_sheet[asin]:
                fails.append(f"xlsx数量不一致 {asin} sheet={smile_sheet[asin]} xlsx={q}")
                print("FAIL: qty mismatch", asin, smile_sheet[asin], q)
            else:
                print("OK  : xlsx qty matches sheet", asin, q)

    print("---")
    if fails:
        print(f"FAILED {len(fails)}")
        return 1
    print("ALL PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
