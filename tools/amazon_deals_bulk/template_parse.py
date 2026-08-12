# -*- coding: utf-8 -*-
"""SC推奨タイムセールxlsxの読取。"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

HERE = Path(__file__).resolve().parent


def load_column_map() -> dict:
    return json.loads((HERE / "column_map.json").read_text(encoding="utf-8"))


def find_template_sheet(wb, candidates: List[str]):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=20, values_only=True):
            vals = [str(c or "") for c in row]
            if any("タイムセールのASIN" in v or "deal_asin" in v for v in vals):
                return ws
    raise RuntimeError("テンプレシートが見つかりません: %s" % wb.sheetnames)


def detect_header(ws, cmap: dict) -> Tuple[int, int, Dict[str, int]]:
    jp_map = cmap.get("jp_to_role") or {}
    best = None
    for r in range(1, 12):
        cells: Dict[str, int] = {}
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            role = jp_map.get(s)
            if not role and "marketplace_id@" in s and "." in s:
                role = s.split(".")[-1]
            if role:
                cells[str(role)] = c
        if "participating" in cells and ("deal_asin" in cells or "deal_price" in cells):
            best = (r, cells)
            break
        if len(cells) >= 6:
            best = (r, cells)
    if not best:
        raise RuntimeError("ヘッダ行を検出できませんでした")
    header_row, colmap = best
    sample = str(ws.cell(header_row, colmap.get("sku", 1)).value or "")
    if sample.startswith("marketplace_id@"):
        data_start = header_row + 1
    else:
        nxt = header_row + 1
        nxt_val = str(ws.cell(nxt, colmap.get("sku", 1)).value or "")
        data_start = nxt + 1 if nxt_val.startswith("marketplace_id@") else nxt
    return header_row, data_start, colmap


def parse_schedule_dates(schedule: str) -> Tuple[Optional[str], Optional[str]]:
    s = str(schedule or "").strip()
    m = re.search(r"\((\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})\)", s)
    if m:
        return m.group(1), m.group(2)
    return None, None


def _parse_date(s: Optional[str]) -> Optional[date]:
    if not s or str(s).upper() == "NA":
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def read_template_rows(path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, int], Any]:
    cmap = load_column_map()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_template_sheet(wb, list(cmap.get("template_sheet_name_candidates") or []))
    _hr, data_start, colmap = detect_header(ws, cmap)
    rows: List[Dict[str, Any]] = []
    for r in range(data_start, (ws.max_row or data_start) + 1):
        asin_c = colmap.get("deal_asin")
        if not asin_c:
            break
        asin = str(ws.cell(r, asin_c).value or "").strip()
        if not asin:
            if r > data_start + 3:
                empty = not any(ws.cell(r, c).value for c in colmap.values())
                if empty:
                    break
            continue
        item = {"_row": r}
        for role, c in colmap.items():
            item[role] = ws.cell(r, c).value
        rows.append(item)
    return rows, colmap, wb


def collect_schedules(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """行データからスケジュール候補を日付順で返す（deal_type / participating 付き）。"""
    seen = set()
    out: List[Dict[str, Any]] = []
    for row in rows:
        sched = str(row.get("schedule") or "").strip()
        if not sched or sched in seen:
            continue
        seen.add(sched)
        start = str(row.get("start_date") or "").strip()
        end = str(row.get("end_date") or "").strip()
        if start.upper() == "NA":
            start = ""
        if end.upper() == "NA":
            end = ""
        ps, pe = parse_schedule_dates(sched)
        start = start or (ps or "")
        end = end or (pe or "")
        out.append(
            {
                "schedule": sched,
                "start": start,
                "end": end,
                "deal_type": row.get("deal_type"),
                "participating": row.get("participating"),
                "deal_price": row.get("deal_price"),
                "committed_units": row.get("committed_units"),
                "seller_price": row.get("seller_price"),
                "source": "cell",
            }
        )
    out.sort(key=lambda x: _parse_date(x.get("start")) or _parse_date(x.get("end")) or date.max)
    return out


def _sqref_covers_cell(sqref: object, col_letter: str, row: int) -> bool:
    """dataValidation.sqref が当該セルを含むか。"""
    from openpyxl.utils import column_index_from_string, range_boundaries

    target_col = column_index_from_string(col_letter)
    for part in str(sqref or "").replace(",", " ").split():
        part = part.strip()
        if not part:
            continue
        if "!" in part:
            part = part.split("!")[-1]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(part)
        except Exception:
            continue
        if min_col <= target_col <= max_col and min_row <= row <= max_row:
            return True
    return False


def values_from_defined_name(wb, name: str) -> List[str]:
    """定義名（DropdownRange_N）→セル値リスト。"""
    name = str(name or "").strip().lstrip("=")
    defn = None
    try:
        defn = wb.defined_names[name]
    except Exception:
        try:
            defn = wb.defined_names.get(name)
        except Exception:
            defn = None
    if defn is None:
        return []
    out: List[str] = []
    dests = []
    try:
        dests = list(defn.destinations)
    except Exception:
        dests = []
    if not dests and getattr(defn, "attr_text", None):
        text = str(defn.attr_text)
        if "!" in text:
            sheet_name, coord = text.split("!", 1)
            dests = [(sheet_name.strip("'"), coord)]
    for sheet_name, coord in dests:
        if sheet_name not in getattr(wb, "sheetnames", []):
            continue
        ws = wb[sheet_name]
        try:
            cells = ws[coord]
        except Exception:
            continue
        if hasattr(cells, "value"):
            seq = [[cells]]
        elif isinstance(cells, tuple) and cells and hasattr(cells[0], "value"):
            seq = [cells]
        else:
            seq = cells
        for row in seq:
            if not isinstance(row, (list, tuple)):
                row = (row,)
            for cell in row:
                v = getattr(cell, "value", cell)
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    out.append(s)
    return out


def schedule_dropdown_values_for_cell(wb, ws, *, row: int, col: int) -> List[str]:
    """テンプレ行のスケジュール列データ検証（ドロップダウン）の選択肢。"""
    from openpyxl.utils import get_column_letter

    if not row or not col:
        return []
    letter = get_column_letter(col)
    try:
        dvs = list(ws.data_validations.dataValidation or [])
    except Exception:
        return []
    for dv in dvs:
        if str(getattr(dv, "type", "") or "") != "list":
            continue
        if not _sqref_covers_cell(dv.sqref, letter, int(row)):
            continue
        f1 = str(getattr(dv, "formula1", "") or "").strip().lstrip("=")
        if not f1:
            continue
        if f1.startswith("DropdownRange_"):
            return values_from_defined_name(wb, f1)
        if "!" in f1 and "ValidationDataSheet" in f1:
            sheet_name, coord = f1.split("!", 1)
            sheet_name = sheet_name.strip("'")
            out: List[str] = []
            if sheet_name in wb.sheetnames:
                block = wb[sheet_name][coord]
                if hasattr(block, "value"):
                    block = ((block,),)
                elif isinstance(block, tuple) and block and hasattr(block[0], "value"):
                    block = (block,)
                for cell_row in block:
                    if not isinstance(cell_row, (list, tuple)):
                        cell_row = (cell_row,)
                    for cell in cell_row:
                        v = getattr(cell, "value", None)
                        if v is not None and str(v).strip():
                            out.append(str(v).strip())
            return out
        got = values_from_defined_name(wb, f1)
        if got:
            return got
        return [f1]
    return []


def collect_dated_dropdown_schedules(
    wb,
    ws,
    rows: List[Dict[str, Any]],
    *,
    schedule_col: int,
    osusume_only: bool = True,
) -> List[Dict[str, Any]]:
    """
    おすすめ行のスケジュール列ドロップダウンから、日付付きセールを全て収集。
    セル未選択の候補も含む（当該SKUに提示されている選択肢が正）。
    """
    from schedule_class import is_mainline_osusume_type

    seen = set()
    out: List[Dict[str, Any]] = []
    for row in rows:
        if osusume_only and not is_mainline_osusume_type(row.get("deal_type")):
            continue
        excel_row = int(row.get("_row") or 0)
        if not excel_row or not schedule_col:
            continue
        for name in schedule_dropdown_values_for_cell(
            wb, ws, row=excel_row, col=int(schedule_col)
        ):
            name = str(name or "").strip()
            if not name or name in ("はい", "いいえ") or name in seen:
                continue
            ps, pe = parse_schedule_dates(name)
            if not ps or not pe:
                continue
            seen.add(name)
            out.append(
                {
                    "schedule": name,
                    "start": ps,
                    "end": pe,
                    "deal_type": row.get("deal_type") or "おすすめタイムセール",
                    "participating": row.get("participating"),
                    "deal_price": row.get("deal_price"),
                    "committed_units": row.get("committed_units"),
                    "seller_price": row.get("seller_price"),
                    "source": "dropdown",
                }
            )
    out.sort(key=lambda x: _parse_date(x.get("start")) or date.max)
    return out


def _sched_entry(name: str, start: str = "", end: str = "") -> Dict[str, Any]:
    ps, pe = parse_schedule_dates(name)
    start = start if start and start.upper() != "NA" else (ps or "")
    end = end if end and end.upper() != "NA" else (pe or "")
    return {"schedule": name, "start": start or "", "end": end or ""}


def collect_schedule_catalog(wb) -> List[Dict[str, Any]]:
    """
    ValidationDataSheet＋テンプレ埋込JSONから候補暦を集める。
    行に未選択でも Smile／BF／次の月枠など日付付き公式を提案できる。
    """
    by_name: Dict[str, Dict[str, Any]] = {}

    def upsert(name: str, start: str = "", end: str = "") -> None:
        name = str(name or "").strip()
        if not name or name in ("はい", "いいえ"):
            return
        # スケジュールらしきものだけ
        if not (
            "(" in name
            or name.startswith("月")
            or name.startswith("カスタム")
            or "プライム" in name
            or "スマイル" in name
            or "Smile" in name
            or "ブラック" in name
        ):
            return
        ent = _sched_entry(name, start, end)
        prev = by_name.get(name)
        if not prev:
            by_name[name] = ent
            return
        # 日付が付いた方を優先
        if (not prev.get("start")) and ent.get("start"):
            by_name[name] = ent

    if "ValidationDataSheet" in getattr(wb, "sheetnames", []):
        ws = wb["ValidationDataSheet"]
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(40, ws.max_row or 1),
            max_col=min(20, ws.max_column or 1),
            values_only=True,
        ):
            for v in row:
                if v is None:
                    continue
                upsert(str(v).strip())

    # 埋込 meta: schedules:{'A1VC...-Amazon Smile...':{start_date:"...",end_date:"..."
    for name in getattr(wb, "sheetnames", []) or []:
        if "テンプレート" not in name and "template" not in name.lower():
            continue
        blob = str(wb[name].cell(1, 1).value or "")
        if "schedules:" not in blob and "start_date" not in blob:
            continue
        for m in re.finditer(
            r"'[^']*?-([^']+)':\{([^}]{0,300})\}",
            blob,
        ):
            label = m.group(1).strip()
            body = m.group(2)
            sd = re.search(r'start_date\s*:\s*"(\d{4}-\d{2}-\d{2})"', body)
            ed = re.search(r'end_date\s*:\s*"(\d{4}-\d{2}-\d{2})"', body)
            upsert(label, sd.group(1) if sd else "", ed.group(1) if ed else "")
        break

    out = list(by_name.values())
    out.sort(key=lambda x: _parse_date(x.get("start")) or _parse_date(x.get("end")) or date.max)
    return out


def merge_schedules(*lists: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """同名は日付付きを優先してマージ。"""
    by_name: Dict[str, Dict[str, Any]] = {}
    for lst in lists:
        for c in lst or []:
            name = str(c.get("schedule") or "").strip()
            if not name:
                continue
            prev = by_name.get(name)
            if not prev:
                by_name[name] = dict(c)
                continue
            if (not prev.get("start")) and c.get("start"):
                by_name[name] = dict(c)
            elif (not prev.get("end")) and c.get("end"):
                prev["end"] = c.get("end")
    out = list(by_name.values())
    out.sort(key=lambda x: _parse_date(x.get("start")) or _parse_date(x.get("end")) or date.max)
    return out


def pick_next_schedules(
    candidates: List[Dict[str, Any]],
    *,
    used: set,
    limit: int = 2,
    today: Optional[date] = None,
) -> List[Dict[str, Any]]:
    today = today or date.today()
    picked = []
    for c in candidates:
        if c["schedule"] in used:
            continue
        end_d = _parse_date(c.get("end"))
        start_d = _parse_date(c.get("start"))
        # 終了日が過去ならスキップ。日付不明（イベント名のみ）は候補に残す
        if end_d and end_d < today:
            continue
        if start_d and end_d is None and start_d < today:
            continue
        picked.append(c)
        if len(picked) >= limit:
            break
    return picked
