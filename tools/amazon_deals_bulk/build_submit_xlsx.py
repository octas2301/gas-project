# -*- coding: utf-8 -*-
"""
公式タイムセール提出用xlsx（P1b）。

**targets-first**: シートの B提出対象ごとに1行作る（イベント複数＝複数行）。
既定は成功UL（DEALS_…_114817_submitv2）準拠:
  対象行のみ残す / 開始終了は YYYY-MM-DD 直書き / シート保護OFF / DV再構築。
**②に当該SKUの候補として存在する枠だけ**を書き出す。
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

HERE = Path(__file__).resolve().parent
LOG = logging.getLogger("amazon_deals_bulk.build")

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)

from paths import folder_path, latest_xlsx, load_config  # noqa: E402
from schedule_class import format_ymd, parse_ymd  # noqa: E402
from sheet_schema import LANE_B, SALE_SHEET  # noqa: E402
from sheets_io import read_sheet_rows, sheets_service  # noqa: E402
from template_parse import (  # noqa: E402
    detect_header,
    find_template_sheet,
    load_column_map,
    parse_schedule_dates,
    read_template_rows,
)


def _truthy(v: Any) -> bool:
    return str(v or "").strip().upper() in ("TRUE", "はい", "YES", "Y", "1", "○", "参加")


def _to_float(v: Any) -> Optional[float]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(re.sub(r"[^\d.]+", "", str(v)))
    except ValueError:
        return None


def is_mainline_deal_type(deal_type: Any) -> bool:
    s = str(deal_type or "").strip()
    if not s:
        return False
    if "数量限定" in s or "LIGHTNING" in s.upper():
        return False
    if "おすすめ" in s or "BEST_DEAL" in s.upper() or "BEST DEAL" in s.upper():
        return True
    return False


def event_match(schedule: str, event: str) -> bool:
    """--event フィルタ。all なら常にTrue。"""
    ev = str(event or "all").strip().lower()
    if ev in ("", "all", "*", "全て", "すべて"):
        return True
    s = str(schedule or "")
    low = s.lower()
    aliases = {
        "smile": ("smile", "スマイル"),
        "bf": ("ブラック", "black", "フライデー", "bf", "サイバー", "cyber"),
        "blackfriday": ("ブラック", "black", "フライデー"),
        "prime": ("プライム感謝", "プライムデー", "プライム"),
    }
    keys = aliases.get(ev)
    if keys:
        return any(k.lower() in low or k in s for k in keys)
    return ev in low or event in s


def load_b_targets(
    cfg: dict, *, event: str = "all", sku: str = "", asin: str = ""
) -> List[Dict[str, Any]]:
    svc = sheets_service(write=False)
    sid = str(cfg.get("ads_spreadsheet_id") or "").strip()
    _h, rows = read_sheet_rows(svc, sid, SALE_SHEET)
    sku_f = str(sku or "").strip()
    asin_f = str(asin or "").strip().upper()
    out = []
    for r in rows:
        if str(r.get("レーン") or "").strip() != LANE_B:
            continue
        if sku_f and str(r.get("SKU") or "").strip() != sku_f:
            continue
        if asin_f and str(r.get("ASIN") or "").strip().upper() != asin_f:
            continue
        if not _truthy(r.get("提出対象")):
            continue
        sched = str(r.get("スケジュール") or "")
        if not sched.strip():
            continue
        # 名付き公式・②由来カスタムとも可（syncが提出対象を付けた行）
        if not event_match(sched, event):
            continue
        st = str(r.get("状態") or "").strip()
        if st in ("見送り", "終了", "失敗", "停止"):
            continue
        # UL済／アップロード済でも再作成可（提出対象=はいなら載せる）
        if str(r.get("有効") or "TRUE") and not _truthy(r.get("有効")):
            if str(r.get("有効") or "").strip() != "":
                continue
        out.append(r)
    return out


def mark_targets_uploaded_(cfg: dict, audits: List[dict], *, run_id: str) -> int:
    """提出xlsx作成後: 対象行の状態を UL済（アップロード済み扱い）に更新。提出対象は残す（再作成可）。"""
    from sheets_io import update_row_fields

    sid = str(cfg.get("ads_spreadsheet_id") or "").strip()
    if not sid:
        return 0
    svc = sheets_service(write=True)
    headers, rows = read_sheet_rows(svc, sid, SALE_SHEET)
    if not headers or "状態" not in headers:
        return 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    n = 0
    for a in audits:
        if a.get("action") != "opt_in":
            continue
        sku = str(a.get("sku") or "").strip()
        sch = str(a.get("schedule") or "").strip()
        for i, r in enumerate(rows):
            if str(r.get("SKU") or "").strip() != sku:
                continue
            if str(r.get("スケジュール") or "").strip() != sch:
                continue
            # シート行はヘッダ行の次が1 → sheets API の行番号は i+2
            update_row_fields(
                svc,
                sid,
                SALE_SHEET,
                i + 2,
                {
                    "状態": "UL済",
                    "更新日時": now,
                    "runId": run_id,
                    "メッセージ": f"提出xlsx作成済（アップロード済み扱い） run={run_id}",
                },
            )
            n += 1
            break
    return n


def unmerge_data_area(ws, data_start: int, *, copy_values: bool = True) -> int:
    targets = [m for m in list(ws.merged_cells.ranges) if m.max_row >= data_start]
    n = 0
    for m in targets:
        tl = ws.cell(m.min_row, m.min_col).value
        ref = str(m)
        ws.unmerge_cells(ref)
        if copy_values:
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    cell = ws.cell(r, c)
                    if cell.value is None and tl is not None:
                        cell.value = tl
        n += 1
    return n


def drop_merges_beyond_row(ws, last_row: int) -> int:
    """削除後に残った高行番号の結合を落とす（openpyxlは行削除で結合が残ることがある）。"""
    n = 0
    for m in list(ws.merged_cells.ranges):
        if m.min_row > last_row or m.max_row > last_row:
            ws.unmerge_cells(str(m))
            n += 1
    return n


def assert_no_vertical_data_merges(ws, data_start: int) -> List[str]:
    bad: List[str] = []
    for m in ws.merged_cells.ranges:
        if m.max_row < data_start:
            continue
        if m.max_row > m.min_row:
            bad.append(str(m))
    return bad


def clear_sheet_comments(ws) -> int:
    n = 0
    try:
        for row in ws.iter_rows(max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            for cell in row:
                if getattr(cell, "comment", None) is not None:
                    cell.comment = None
                    n += 1
    except TypeError:
        # openpyxl 環境差で comment 操作が失敗してもUL自体には必須ではない
        LOG.warning("コメント削除をスキップ")
    return n


def disable_template_protection(ws) -> None:
    """成功submitv2: タイムセール推奨テンプレートのシート保護はOFF。"""
    try:
        ws.protection.sheet = False
    except Exception as e:
        LOG.warning("シート保護解除に失敗: %s", e)


def collect_data_rows(ws, data_start: int, colmap: Dict[str, int]) -> List[int]:
    asin_c = colmap.get("deal_asin")
    if not asin_c:
        return []
    rows: List[int] = []
    max_r = ws.max_row or data_start
    for r in range(data_start, max_r + 1):
        asin = str(ws.cell(r, asin_c).value or "").strip()
        if not asin:
            if r > data_start + 5 and not any(
                ws.cell(r, c).value for c in colmap.values() if c
            ):
                break
            continue
        rows.append(r)
    return rows


def delete_rows_desc(ws, row_nums: List[int]) -> int:
    n = 0
    for r in sorted(set(row_nums), reverse=True):
        ws.delete_rows(r, 1)
        n += 1
    return n


def snapshot_row(ws, r: int, max_col: int) -> List[Any]:
    return [ws.cell(r, c).value for c in range(1, max_col + 1)]


def write_row_values(ws, r: int, values: List[Any]) -> None:
    for c, v in enumerate(values, start=1):
        ws.cell(r, c).value = v


def find_prototype(
    ws,
    data_rows: List[int],
    colmap: Dict[str, int],
    *,
    sku: str,
    asin: str,
    schedule: str = "",
) -> Optional[int]:
    """おすすめ行をスケジュール一致→SKU→ASIN優先で探す。"""
    type_c = colmap.get("deal_type")
    sku_c = colmap.get("sku")
    asin_c = colmap.get("deal_asin")
    sched_c = colmap.get("schedule")
    by_sku = None
    by_asin = None
    want = str(schedule or "").strip()
    for r in data_rows:
        dt = str(ws.cell(r, type_c).value or "") if type_c else ""
        if not is_mainline_deal_type(dt):
            continue
        rsku = str(ws.cell(r, sku_c).value or "").strip() if sku_c else ""
        rasin = str(ws.cell(r, asin_c).value or "").strip().upper() if asin_c else ""
        rsched = str(ws.cell(r, sched_c).value or "").strip() if sched_c else ""
        if want and rsku == sku and rsched == want:
            return r
        if sku and rsku == sku and by_sku is None:
            by_sku = r
        if asin and rasin == asin and by_asin is None:
            by_asin = r
    return by_sku or by_asin


def load_sku_schedules_from_source(src: Path) -> Dict[str, set]:
    """SKU → ②セル＋スケジュール列ドロップダウン（日付付き）のスケジュール名集合。"""
    from template_parse import collect_dated_dropdown_schedules

    rows, colmap, wb = read_template_rows(src)
    cmap = load_column_map()
    try:
        ws = find_template_sheet(wb, list(cmap.get("template_sheet_name_candidates") or []))
    except Exception:
        ws = None
    schedule_col = int(colmap.get("schedule") or 0)
    out: Dict[str, set] = {}
    by_sku: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        sku = str(r.get("sku") or "").strip()
        if not sku:
            continue
        by_sku.setdefault(sku, []).append(r)
        sched = str(r.get("schedule") or "").strip()
        if sched:
            out.setdefault(sku, set()).add(sched)
    if ws is not None and schedule_col:
        for sku, cands in by_sku.items():
            for d in collect_dated_dropdown_schedules(
                wb, ws, cands, schedule_col=schedule_col, osusume_only=True
            ):
                name = str(d.get("schedule") or "").strip()
                if name:
                    out.setdefault(sku, set()).add(name)
    return out


def _set_text_ymd_cell(cell, value: Any) -> str:
    """
    成功UL(submitv2)準拠: 開始/終了は YYYY-MM-DD 文字列＋書式 yyyy\\-mm\\-dd。
    （VLOOKUP残しや @ 書式は SC「ファイル処理失敗」の原因になり得る）
    """
    ymd = format_ymd(value)
    if not ymd:
        return ""
    cell.value = ymd
    cell.number_format = r"yyyy\-mm\-dd"
    return ymd


def _set_text_cell(cell, value: str) -> None:
    cell.value = str(value or "")
    cell.number_format = "General"


def resolve_submit_dates(erow: Dict[str, Any]) -> tuple:
    """
    台帳の開始/終了を YYYY-MM-DD に。空ならスケジュール名の (YYYY-MM-DD - YYYY-MM-DD) から補完。
    戻り値: (start_ymd, end_ymd, err_msg)
    """
    sched = str(erow.get("スケジュール") or "").strip()
    start = format_ymd(erow.get("開始日"))
    end = format_ymd(erow.get("終了日"))
    ps, pe = parse_schedule_dates(sched)
    if not start and ps:
        start = ps
    if not end and pe:
        end = pe
    if not start or not end:
        return "", "", "開始日/終了日が YYYY-MM-DD で取れない"
    d0, d1 = parse_ymd(start), parse_ymd(end)
    if not d0 or not d1:
        return "", "", "開始日/終了日の解析失敗"
    if d1 < d0:
        return "", "", "終了日が開始日より前"
    span = (d1 - d0).days + 1
    if span < 1 or span > 14:
        return "", "", "期間は1〜14日（データ定義）got=%s" % span
    # スケジュール名に日付があるときはその枠内に収める
    if ps and pe:
        w0, w1 = parse_ymd(ps), parse_ymd(pe)
        if w0 and w1 and (d0 < w0 or d1 > w1):
            return "", "", "開始/終了がスケジュール枠外 (%s..%s vs %s..%s)" % (start, end, ps, pe)
    return start, end, ""


def rebuild_compact_validations(
    wb,
    ws,
    colmap: Dict[str, int],
    *,
    data_start: int,
    n_rows: int,
    schedule_options: List[str],
) -> None:
    """
    成功submitv2相当: 参加中・スケジュールのDVを残り行だけに張り直す。
    ValidationDataSheet も最小構成に作り直す。
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    ws.data_validations.dataValidation = []
    for name in list(wb.defined_names):
        if str(name).startswith("DropdownRange_"):
            del wb.defined_names[name]

    if "ValidationDataSheet" not in wb.sheetnames:
        wb.create_sheet("ValidationDataSheet")
    vd = wb["ValidationDataSheet"]
    if vd.max_row and vd.max_row > 0:
        vd.delete_rows(1, vd.max_row)
    if vd.max_column and vd.max_column > 1:
        vd.delete_cols(2, vd.max_column - 1)
    vd["A1"] = "はい"
    vd["A2"] = "いいえ"
    opts = []
    seen = set()
    for o in schedule_options:
        s = str(o or "").strip()
        if s and s not in seen:
            seen.add(s)
            opts.append(s)
    if not opts:
        opts = ["NA"]
    for i, o in enumerate(opts, start=1):
        vd.cell(i, 2).value = o

    wb.defined_names.add(
        DefinedName(name="DropdownRange_0", attr_text="ValidationDataSheet!$A$1:$A$2")
    )
    wb.defined_names.add(
        DefinedName(
            name="DropdownRange_1",
            attr_text="ValidationDataSheet!$B$1:$B$%s" % len(opts),
        )
    )

    end_row = data_start + n_rows - 1
    part_l = get_column_letter(int(colmap["participating"]))
    sched_l = get_column_letter(int(colmap["schedule"]))
    dv_part = DataValidation(type="list", formula1="DropdownRange_0", allow_blank=True)
    dv_part.add("%s%s:%s%s" % (part_l, data_start, part_l, end_row))
    ws.add_data_validation(dv_part)
    dv_sch = DataValidation(type="list", formula1="DropdownRange_1", allow_blank=True)
    for r in range(data_start, end_row + 1):
        dv_sch.add("%s%s" % (sched_l, r))
    ws.add_data_validation(dv_sch)


def apply_target_to_row(ws, r: int, colmap: Dict[str, int], erow: Dict[str, Any]) -> dict:
    """
    成功UL(DEALS_…_114817_submitv2)準拠:
    参加中・スケジュール・開始/終了（文字）・価格・数量を書く。
    開始/終了に VLOOKUP を残さない（全行残し＋unmergeだと参照壊れでSC処理失敗）。
    """
    audit: Dict[str, Any] = {"row": r, "action": "opt_in"}

    def cell(role: str):
        c = colmap.get(role)
        return ws.cell(r, c) if c else None

    audit["asin"] = str(erow.get("ASIN") or "").strip().upper()
    audit["sku"] = str(erow.get("SKU") or "").strip()
    if cell("participating") is not None:
        _set_text_cell(cell("participating"), "はい")
    sched = str(erow.get("スケジュール") or "").strip()
    if cell("schedule") is not None and sched:
        _set_text_cell(cell("schedule"), sched)
    start_ymd, end_ymd, derr = resolve_submit_dates(erow)
    if derr:
        audit["action"] = "miss"
        audit["reason"] = derr
        audit["schedule"] = sched
        LOG.error(
            "日付不正のため行スキップ: %s | %s | %s",
            audit["sku"],
            sched,
            derr,
        )
        return audit
    if cell("start_date") is not None:
        _set_text_ymd_cell(cell("start_date"), start_ymd)
    if cell("end_date") is not None:
        _set_text_ymd_cell(cell("end_date"), end_ymd)
    dp = cell("deal_price")
    sc_price = _to_float(dp.value) if dp is not None else None
    want = _to_float(erow.get("タイムセール価格_確定"))
    if dp is not None and want is not None:
        if sc_price is not None and want > sc_price + 1e-9:
            want = sc_price
        dp.value = want
    qty_c = cell("committed_units")
    want_q = _to_float(erow.get("販売商品数_確定"))
    if qty_c is not None and want_q is not None:
        qty_c.value = int(want_q)
    audit["schedule"] = sched
    audit["start"] = start_ymd
    audit["end"] = end_ymd
    audit["dates_via"] = "literal_ymd_success_layout"
    audit["qty"] = int(want_q) if want_q is not None else None
    return audit


def build(
    *,
    cfg: dict,
    source: Optional[Path],
    write: bool,
    inspect_only: bool,
    event: str = "all",
    sku: str = "",
    asin: str = "",
    only_schedule: str = "",
    in_place: bool = False,
) -> int:
    folder02 = folder_path(cfg, "02")
    folder03 = folder_path(cfg, "03")
    src = source or latest_xlsx(folder02)
    if not src or not src.is_file():
        LOG.error("②にxlsxがありません: %s", folder02)
        return 1

    # 既定=成功submitv2形（対象行のみ・日付直書き・保護OFF）。--in-place は非推奨
    compact = not bool(in_place)
    LOG.info(
        "入力: %s event=%s sku=%s asin=%s only_schedule=%s compact=%s",
        src,
        event,
        sku or "-",
        asin or "-",
        only_schedule or "-",
        compact,
    )
    cmap = load_column_map()
    wb = openpyxl.load_workbook(src)
    ws = find_template_sheet(wb, list(cmap.get("template_sheet_name_candidates") or []))
    header_row, data_start, colmap = detect_header(ws, cmap)
    LOG.info("sheet=%s header=%s data=%s", ws.title, header_row, data_start)

    if inspect_only:
        return 0

    targets = load_b_targets(cfg, event=event, sku=sku, asin=asin)
    want_sched = str(only_schedule or "").strip()
    if want_sched:
        targets = [
            t
            for t in targets
            if str(t.get("スケジュール") or "").strip() == want_sched
        ]
    sku_schedules = load_sku_schedules_from_source(src)
    filtered: List[Dict[str, Any]] = []
    for t in targets:
        tsku = str(t.get("SKU") or "").strip()
        tsched = str(t.get("スケジュール") or "").strip()
        allowed = sku_schedules.get(tsku) or set()
        if tsched not in allowed:
            LOG.warning(
                "②ドロップダウンに無いスケジュールを除外: %s | %s",
                tsku,
                tsched,
            )
            continue
        filtered.append(t)
    targets = filtered
    date_ok: List[Dict[str, Any]] = []
    for t in targets:
        _s, _e, derr = resolve_submit_dates(t)
        if derr:
            LOG.error(
                "提出対象を除外（日付）: %s | %s | %s",
                t.get("SKU"),
                t.get("スケジュール"),
                derr,
            )
            continue
        t = dict(t)
        t["開始日"], t["終了日"] = _s, _e
        date_ok.append(t)
    targets = date_ok
    # 同一ASINは1ファイル1スケジュール（複数枠同時は冷却エラーの主因）
    by_asin: Dict[str, List[Dict[str, Any]]] = {}
    for t in targets:
        a = str(t.get("ASIN") or "").strip().upper()
        by_asin.setdefault(a, []).append(t)
    deduped: List[Dict[str, Any]] = []
    for a, lst in by_asin.items():
        if len(lst) > 1 and not want_sched:
            LOG.warning(
                "同一ASINに複数枠→開始日が最も近い1件のみ: %s %s",
                a,
                [x.get("スケジュール") for x in lst],
            )
            lst = sorted(lst, key=lambda x: format_ymd(x.get("開始日")) or "9999")
            deduped.append(lst[0])
        else:
            deduped.extend(lst)
    targets = deduped
    LOG.info("B提出対象(event=%s)=%s", event, len(targets))
    if not targets:
        LOG.error("提出対象が0件です。event=%s / only_schedule / シートを確認してください。", event)
        return 1

    data_rows = collect_data_rows(ws, data_start, colmap)
    max_col = max(colmap.values()) if colmap else (ws.max_column or 20)
    part_c = colmap.get("participating")
    type_c = colmap.get("deal_type")
    audits: List[dict] = []
    deleted = 0
    mode_name = "compact_success_layout" if compact else "in_place_opt_in_discouraged"

    if compact:
        planned: List[Tuple[List[Any], Dict[str, Any]]] = []
        for t in targets:
            tsku = str(t.get("SKU") or "").strip()
            tasin = str(t.get("ASIN") or "").strip().upper()
            proto = find_prototype(
                ws,
                data_rows,
                colmap,
                sku=tsku,
                asin=tasin,
                schedule=str(t.get("スケジュール") or ""),
            )
            if not proto:
                audits.append(
                    {
                        "action": "miss",
                        "sku": tsku,
                        "asin": tasin,
                        "schedule": t.get("スケジュール"),
                        "reason": "②テンプレにおすすめ行なし",
                    }
                )
                continue
            planned.append((snapshot_row(ws, proto, max_col), t))
        if not planned:
            LOG.error("書き込める行がありません")
            return 1
        # 行削除前にデータ結合を解除（値コピーなし＝VLOOKUP参照壊れを避ける）
        unmerged = unmerge_data_area(ws, data_start, copy_values=False)
        LOG.info("compact前の結合解除=%s", unmerged)
        deleted = delete_rows_desc(ws, data_rows)
        LOG.info(
            "compact_success: 非対象行削除=%s（成功submitv2と同型: 対象行のみ）",
            deleted,
        )
        dropped = drop_merges_beyond_row(ws, data_start + len(planned) - 1)
        if dropped:
            LOG.info("余剰結合削除=%s", dropped)
        n_comments = clear_sheet_comments(ws)
        if n_comments:
            LOG.info("コメント削除=%s", n_comments)
        for i, (vals, t) in enumerate(planned):
            r = data_start + i
            write_row_values(ws, r, vals)
            audits.append(apply_target_to_row(ws, r, colmap, t))
        sched_opts = [str(t.get("スケジュール") or "").strip() for _, t in planned]
        # 当該SKUのドロップダウン候補も選択肢に残す（成功v2も複数候補列があった）
        for _, t in planned:
            tsku = str(t.get("SKU") or "").strip()
            sched_opts.extend(sorted(sku_schedules.get(tsku) or []))
        rebuild_compact_validations(
            wb,
            ws,
            colmap,
            data_start=data_start,
            n_rows=len(planned),
            schedule_options=sched_opts,
        )
        disable_template_protection(ws)
        LOG.info(
            "compact_success: 参加はい=%s / 保護OFF / DV再構築 / 日付=直書き",
            sum(1 for a in audits if a.get("action") == "opt_in"),
        )
    else:
        LOG.warning(
            "--in-place: 全行残し。unmergeでVLOOKUP参照が壊れ SC処理失敗しやすい。非推奨"
        )
        unmerged = unmerge_data_area(ws, data_start)
        LOG.info("データ領域の結合解除=%s", unmerged)
        if part_c:
            for r in data_rows:
                dt = str(ws.cell(r, type_c).value or "") if type_c else ""
                if is_mainline_deal_type(dt) or "数量限定" in dt:
                    _set_text_cell(ws.cell(r, part_c), "いいえ")
        opt_rows: Dict[str, int] = {}
        for t in targets:
            tsku = str(t.get("SKU") or "").strip()
            tasin = str(t.get("ASIN") or "").strip().upper()
            proto = find_prototype(
                ws,
                data_rows,
                colmap,
                sku=tsku,
                asin=tasin,
                schedule="",
            )
            if not proto:
                audits.append(
                    {
                        "action": "miss",
                        "sku": tsku,
                        "asin": tasin,
                        "schedule": t.get("スケジュール"),
                        "reason": "②テンプレにおすすめ行なし",
                    }
                )
                LOG.warning("原型なし: %s %s", tsku, tasin)
                continue
            if tsku in opt_rows:
                LOG.warning("同一SKUが重複ターゲット→後勝ち: %s", tsku)
            opt_rows[tsku] = proto
            audits.append(apply_target_to_row(ws, proto, colmap, t))
        disable_template_protection(ws)

    bad_merges = assert_no_vertical_data_merges(ws, data_start)
    if bad_merges:
        LOG.error("データ行に縦結合が残っています: %s", bad_merges[:10])
        return 1

    run_id = datetime.now().strftime("DEALS_%Y%m%d_%H%M%S")
    if want_sched:
        ev_slug = re.sub(r"[^\w]+", "", want_sched)[:40] or "sched"
    else:
        ev_slug = re.sub(r"[^\w]+", "", event)[:12] or "all"
    opt_in_n = sum(1 for a in audits if a.get("action") == "opt_in")
    summary = {
        "run_id": run_id,
        "source": str(src),
        "event": event,
        "only_schedule": want_sched,
        "opt_in": opt_in_n,
        "miss": sum(1 for a in audits if a.get("action") == "miss"),
        "rows_deleted": deleted,
        "mode": mode_name,
        "deal_type_policy": "対象SKU行のみ残す（成功submitv2準拠）"
        if compact
        else "おすすめ残し・他はいえ",
        "note": "成功UL差分準拠: compact+日付直書き(yyyy-mm-dd)+保護OFF+DV再構築",
        "write": write,
    }
    LOG.info("summary=%s", summary)

    if not write:
        (HERE / "_work").mkdir(exist_ok=True)
        (HERE / "_work" / f"{run_id}_audit.json").write_text(
            json.dumps({"summary": summary, "audits": audits}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        LOG.info("dry-run。本番は --write")
        return 0

    folder03.mkdir(parents=True, exist_ok=True)
    out_path = folder03 / f"{run_id}_{ev_slug}_submit.xlsx"
    wb.save(out_path)
    (folder03 / f"{run_id}_{ev_slug}_audit.json").write_text(
        json.dumps({"summary": summary, "audits": audits}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    LOG.info("出力: %s（参加はい=%s）", out_path, opt_in_n)
    try:
        n_mark = mark_targets_uploaded_(cfg, audits, run_id=run_id)
        LOG.info("シート状態→UL済（作成=アップロード済み扱い）: %s行", n_mark)
    except Exception as e:
        LOG.warning("シートUL済更新に失敗: %s", e)
    return 0 if opt_in_n else 1


def archive_older_02_files(cfg: dict, *, keep_latest: int = 1) -> List[str]:
    """②の古いxlsxを『使用済み』サブフォルダへ移動。"""
    import shutil

    folder02 = folder_path(cfg, "02")
    if not folder02.is_dir():
        return []
    used = folder02 / "使用済み"
    used.mkdir(exist_ok=True)
    files = [
        p
        for p in folder02.iterdir()
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")
    ]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    moved = []
    for p in files[keep_latest:]:
        dest = used / p.name
        if dest.exists():
            dest = used / ("%s_%s" % (p.stem, datetime.now().strftime("%H%M%S")) + p.suffix)
        shutil.move(str(p), str(dest))
        moved.append(str(dest))
        LOG.info("②使用済みへ移動: %s", dest.name)
    return moved


def main(argv=None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser(description="Amazon公式B提出xlsx (targets-first)")
    ap.add_argument("--config", type=Path, default=None)
    ap.add_argument("--source", type=Path, default=None)
    ap.add_argument("--inspect", action="store_true")
    ap.add_argument("--write", action="store_true")
    ap.add_argument(
        "--event",
        type=str,
        default="all",
        help="all / smile / bf / prime （シートBのスケジュールで絞る）",
    )
    ap.add_argument("--sku", type=str, default="", help="1SKUだけ提出xlsx（テスト用）")
    ap.add_argument("--asin", type=str, default="", help="1ASINだけ提出xlsx（テスト用）")
    ap.add_argument(
        "--only-schedule",
        type=str,
        default="",
        help="このスケジュール名と完全一致する台帳行だけ提出（8月/9月の分割用）",
    )
    ap.add_argument(
        "--in-place",
        action="store_true",
        help="非推奨: 全行残し（VLOOKUP壊れでSC処理失敗しやすい）。既定は成功submitv2形のcompact",
    )
    ap.add_argument(
        "--archive-02",
        action="store_true",
        help="②の最新以外を『使用済み』へ移動",
    )
    args = ap.parse_args(argv)
    local = HERE / "config.local.json"
    if not local.is_file():
        local.write_text(
            (HERE / "config.example.json").read_text(encoding="utf-8"), encoding="utf-8"
        )
    cfg = load_config(args.config or local)
    if args.archive_02:
        archive_older_02_files(cfg, keep_latest=1)
    return build(
        cfg=cfg,
        source=args.source,
        write=bool(args.write),
        inspect_only=bool(args.inspect),
        event=str(args.event or "all"),
        sku=str(args.sku or ""),
        asin=str(args.asin or ""),
        only_schedule=str(args.only_schedule or ""),
        in_place=bool(args.in_place),
    )


if __name__ == "__main__":
    raise SystemExit(main())
