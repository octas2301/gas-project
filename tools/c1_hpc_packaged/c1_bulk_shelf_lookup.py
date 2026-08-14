#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B-T1: テンプレ棚引き（PT＋指紋）→ SC DL要否を 05 に出す。

正本: docs/org/LV4_LANE_B_BULK_TEMPLATE_T1_APPROVAL.md
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(2)

from c1_packaged import compute_header_fingerprint  # noqa: E402

SCRIPT_DIR = Path(__file__).resolve().parent
LOG = logging.getLogger("c1_bulk_shelf_lookup")

DEFAULT_DRIVE04 = Path(r"G:/マイドライブ/04.amazonカタログ作成（CSV一括UL）")
DEFAULT_SHELF = DEFAULT_DRIVE04 / "06.純正テンプレ原本（読取専用・触らない）"
DEFAULT_REPORT = DEFAULT_DRIVE04 / "05.SC処理結果・ログ退避（人間）"
DEFAULT_REGISTRY = SCRIPT_DIR / "shelf_registry.json"
SHEET_CANDIDATES = ("テンプレート", "Template", "template")


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _safe_name(name: str) -> str:
    s = re.sub(r"[^\w.\-]+", "_", name, flags=re.UNICODE)
    return s[:80] or "pt"


def _load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _pick_sheet(wb: Any) -> Tuple[str, Any]:
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return name, wb[name]
    for name in wb.sheetnames:
        if "テンプレ" in name or "template" in name.lower():
            return name, wb[name]
    raise RuntimeError("テンプレートシートが見つかりません: %s" % (wb.sheetnames,))


def _entry_for_pt(registry: dict, product_type: str) -> Optional[dict]:
    pt = (product_type or "").strip().upper()
    for e in registry.get("entries") or []:
        if str(e.get("productType") or "").strip().upper() == pt:
            return e
    return None


def _fingerprint_file(
    xlsm_path: Path,
    rows: List[int],
    max_col: int,
) -> str:
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    try:
        _, ws = _pick_sheet(wb)
        return compute_header_fingerprint(ws, rows, max_col)
    finally:
        wb.close()


def lookup_shelf(
    product_type: str,
    shelf_dir: Path,
    registry: dict,
) -> Dict[str, Any]:
    entry = _entry_for_pt(registry, product_type)
    if not entry:
        return {
            "status": "DL_REQUIRED",
            "reason": "registry_miss",
            "productType": product_type,
            "message": "レジストリに Product Type がありません。SCで純正をDLし 09 へ保存してください。",
            "templatePath": None,
            "expectedSha256": None,
            "matchedSha256": None,
        }

    expected = str(entry.get("fingerprintSha256") or "").strip()
    rows = list(entry.get("fingerprintRows") or [3, 4, 5])
    max_col = int(entry.get("fingerprintMaxCol") or 321)
    preferred = str(entry.get("preferredFileName") or "").strip()

    if not shelf_dir.is_dir():
        return {
            "status": "DL_REQUIRED",
            "reason": "shelf_dir_missing",
            "productType": product_type,
            "message": "棚フォルダ（06）がありません。",
            "templatePath": None,
            "expectedSha256": expected,
            "matchedSha256": None,
            "registryEntry": entry,
        }

    candidates: List[Path] = []
    if preferred:
        p = shelf_dir / preferred
        if p.is_file():
            candidates.append(p)
    for p in sorted(shelf_dir.glob("*.xlsm")):
        if p not in candidates:
            candidates.append(p)

    scanned: List[dict] = []
    for path in candidates:
        try:
            sha = _fingerprint_file(path, rows, max_col)
        except Exception as e:
            scanned.append({"path": str(path), "error": str(e)})
            continue
        scanned.append({"path": str(path), "sha256": sha})
        if expected and sha == expected:
            return {
                "status": "DL_NOT_NEEDED",
                "reason": "fingerprint_match",
                "productType": product_type,
                "message": "SCダウンロード不要。06の登録済み純正を使用してください。",
                "templatePath": str(path),
                "expectedSha256": expected,
                "matchedSha256": sha,
                "registryEntry": entry,
                "scanned": scanned,
            }

    return {
        "status": "DL_REQUIRED",
        "reason": "fingerprint_miss" if candidates else "shelf_empty",
        "productType": product_type,
        "message": "棚に一致する純正がありません。SCで当該バルクをDLし 09 へ保存してください。",
        "templatePath": None,
        "expectedSha256": expected,
        "matchedSha256": None,
        "registryEntry": entry,
        "scanned": scanned,
    }


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="B-T1 template shelf lookup (DL need?)")
    parser.add_argument("--product-type", required=True, help="例: SEASONING")
    parser.add_argument("--shelf-dir", default=str(DEFAULT_SHELF))
    parser.add_argument("--registry", default=str(DEFAULT_REGISTRY))
    parser.add_argument("--report-dir", default=str(DEFAULT_REPORT))
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    registry_path = Path(args.registry)
    registry = _load_json(registry_path)
    result = lookup_shelf(args.product_type, Path(args.shelf_dir), registry)
    result["runId"] = "B_T1_SHELF_%s_%s" % (_safe_name(args.product_type), _utc_stamp())
    result["registryPath"] = str(registry_path)
    result["shelfDir"] = str(args.shelf_dir)

    report_dir = Path(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    json_path = report_dir / ("%s.json" % result["runId"])
    txt_path = report_dir / ("%s_SUMMARY.txt" % result["runId"])
    _save_json(json_path, result)
    summary = (
        "runId=%s\nstatus=%s\nreason=%s\nproductType=%s\nmessage=%s\n"
        "templatePath=%s\nexpectedSha=%s\nmatchedSha=%s\n"
        % (
            result["runId"],
            result["status"],
            result.get("reason"),
            result.get("productType"),
            result.get("message"),
            result.get("templatePath"),
            result.get("expectedSha256"),
            result.get("matchedSha256"),
        )
    )
    txt_path.write_text(summary, encoding="utf-8")
    LOG.info("%s", summary.replace("\n", " | "))
    LOG.info("レポート: %s", json_path)

    # DL不要=0 / 要DL=2 / その他=1
    if result["status"] == "DL_NOT_NEEDED":
        return 0
    if result["status"] == "DL_REQUIRED":
        return 2
    return 1


if __name__ == "__main__":
    sys.exit(main())
