#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SHELF Browse 網羅抽出: 純正 xlsm の「データを閲覧する」全行 →
shelf_browse_catalog.json ＋ shelf_registry.json の browseIndex マージ。

正本: docs/org/LV4_SHELF_BROWSE_CATALOG_APPROVAL.md
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です", file=sys.stderr)
    sys.exit(2)

SCRIPT_DIR = Path(__file__).resolve().parent
LOG = logging.getLogger("c1_shelf_browse_extract")
DEFAULT_06 = Path(
    r"G:/マイドライブ/04.amazonカタログ作成（CSV一括UL）/06.純正テンプレ原本（読取専用・触らない）"
)
DEFAULT_DRIVE04 = Path(r"G:/マイドライブ/04.amazonカタログ作成（CSV一括UL）")
SHEET_NAME = "データを閲覧する"
KNOWN_TEMPLATES = {
    "FOOD_FISH_GROCERY.xlsm": {
        "allowedProductTypes": ["FOOD", "FISH", "GROCERY"],
        "preferredProductType": "GROCERY",
        "fingerprintSha": "74ccdcf96c22879dc80cbe87e8b41aa615e923529f151f091552ccbe3cefb010",
        "columnMapPath": "food_fish_grocery_column_map.json",
    },
    "FOOD_HERB_SEASONING_FISH_VEGETABLE.xlsm": {
        "allowedProductTypes": ["VEGETABLE", "FISH", "SAUCE", "SEASONING", "HERB", "FOOD"],
        "preferredProductType": "SEASONING",
        "fingerprintSha": "57190dbc80b8494201e52242874037a4782e80ada4764e53d503564454c2b8e5",
        "columnMapPath": "food_seasoning_column_map.json",
    },
}


def _utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _norm_id(v: Any) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    m = re.search(r"(\d{6,})", s)
    return m.group(1) if m else s


def extract_browse_rows(
    xlsm: Path,
    allowed: List[str],
    preferred: str,
    fingerprint: str,
    column_map: str,
    template_url: str = "",
) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise RuntimeError("シートがありません: %s / %s" % (SHEET_NAME, wb.sheetnames))
    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        raise RuntimeError("ヘッダー空: %s" % xlsm.name)
    bn_i = bp_i = None
    for j, h in enumerate(header):
        hs = str(h or "").strip()
        if hs == "Browse Node" or hs.replace(" ", "").lower() == "browsenode":
            bn_i = j
        if "BrowsePath" in hs.replace(" ", "") or hs == "Browse Path":
            bp_i = j
    if bn_i is None:
        wb.close()
        raise RuntimeError("Browse Node 列がありません: %s" % xlsm.name)
    out: List[Dict[str, Any]] = []
    extracted = _utc()
    for raw in rows_iter:
        if not raw:
            continue
        node = _norm_id(raw[bn_i] if bn_i < len(raw) else "")
        if not node:
            continue
        path = ""
        if bp_i is not None and bp_i < len(raw) and raw[bp_i] is not None:
            path = str(raw[bp_i]).strip()
        out.append(
            {
                "browseNodeId": node,
                "browsePath": path,
                "templateFile": xlsm.name,
                "templateUrl": template_url or "",
                "allowedProductTypes": list(allowed),
                "preferredProductType": preferred,
                "fingerprintSha": fingerprint,
                "columnMapPath": column_map,
                "extractedAt": extracted,
                "sourceSheet": SHEET_NAME,
            }
        )
    wb.close()
    return out


def merge_catalog(
    existing: Dict[str, Any],
    new_rows: List[Dict[str, Any]],
    template_file: str,
) -> Dict[str, Any]:
    """同一 templateFile の旧行を置き換え、他テンプレは維持。"""
    keep = [
        r
        for r in (existing.get("rows") or [])
        if str(r.get("templateFile") or "") != template_file
    ]
    keep.extend(new_rows)
    return {
        "version": "SHELF-BROWSE-1",
        "updatedAt": _utc(),
        "notes": "データを閲覧する 全行網羅。照合キー=browseNodeId。",
        "rows": keep,
    }


def sync_registry_browse_index(registry_path: Path, catalog: Dict[str, Any]) -> None:
    reg = json.loads(registry_path.read_text(encoding="utf-8"))
    reg["version"] = "B-T1-3"
    reg["browseCatalogPath"] = "shelf_browse_catalog.json"
    reg["browseIndex"] = catalog.get("rows") or []
    reg["notes"] = (
        "PT+fingerprintSha256=充填検証。browseIndex=ルーティング（Node ID）。"
        "Catalog MEAT 等は Node 解決後の preferredPT 優先／エイリアスはフォールバック。"
    )
    registry_path.write_text(
        json.dumps(reg, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    ap = argparse.ArgumentParser(description="SHELF Browse 網羅抽出")
    ap.add_argument("--template-dir", default=str(DEFAULT_06))
    ap.add_argument("--xlsm", default="", help="単一 xlsm（省略時は dir 内の既知ファイル）")
    ap.add_argument("--allowed-pts", default="", help="FOOD,FISH,GROCERY")
    ap.add_argument("--preferred-pt", default="")
    ap.add_argument("--fingerprint", default="")
    ap.add_argument("--column-map", default="")
    ap.add_argument("--catalog-out", default=str(SCRIPT_DIR / "shelf_browse_catalog.json"))
    ap.add_argument("--registry", default=str(SCRIPT_DIR / "shelf_registry.json"))
    ap.add_argument("--sync-drive04", action="store_true", default=True)
    ap.add_argument("--no-sync-drive04", action="store_true")
    args = ap.parse_args()

    catalog_path = Path(args.catalog_out)
    existing: Dict[str, Any] = {"rows": []}
    if catalog_path.is_file():
        existing = json.loads(catalog_path.read_text(encoding="utf-8"))

    targets: List[Tuple[Path, dict]] = []
    if args.xlsm:
        p = Path(args.xlsm)
        meta = dict(KNOWN_TEMPLATES.get(p.name) or {})
        if args.allowed_pts:
            meta["allowedProductTypes"] = [x.strip().upper() for x in args.allowed_pts.split(",") if x.strip()]
        if args.preferred_pt:
            meta["preferredProductType"] = args.preferred_pt.strip().upper()
        if args.fingerprint:
            meta["fingerprintSha"] = args.fingerprint.strip()
        if args.column_map:
            meta["columnMapPath"] = args.column_map.strip()
        if not meta.get("allowedProductTypes") or not meta.get("preferredProductType"):
            raise SystemExit("allowed/preferred PT を指定するか KNOWN_TEMPLATES に追加してください")
        targets.append((p, meta))
    else:
        tdir = Path(args.template_dir)
        for name, meta in KNOWN_TEMPLATES.items():
            p = tdir / name
            if p.is_file():
                targets.append((p, meta))
            else:
                LOG.warning("スキップ（無し）: %s", p)

    if not targets:
        raise SystemExit("抽出対象 xlsm がありません")

    catalog = existing
    for path, meta in targets:
        rows = extract_browse_rows(
            path,
            meta["allowedProductTypes"],
            meta["preferredProductType"],
            meta.get("fingerprintSha") or "",
            meta.get("columnMapPath") or "",
        )
        LOG.info("extracted %s rows=%d preferred=%s", path.name, len(rows), meta["preferredProductType"])
        catalog = merge_catalog(catalog, rows, path.name)

    catalog_path.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    LOG.info("wrote %s total_rows=%d", catalog_path, len(catalog.get("rows") or []))

    reg_path = Path(args.registry)
    sync_registry_browse_index(reg_path, catalog)
    LOG.info("updated registry browseIndex: %s", reg_path)

    if not args.no_sync_drive04:
        drive04 = DEFAULT_DRIVE04
        if drive04.is_dir():
            shutil.copy2(catalog_path, drive04 / "shelf_browse_catalog.json")
            shutil.copy2(reg_path, drive04 / "shelf_registry.json")
            LOG.info("synced Drive04 registry+catalog")

    print(
        json.dumps(
            {
                "ok": True,
                "totalRows": len(catalog.get("rows") or []),
                "templates": sorted({r["templateFile"] for r in catalog.get("rows") or []}),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
