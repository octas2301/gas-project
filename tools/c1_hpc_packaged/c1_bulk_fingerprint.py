#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B-T0: SC純正バルク xlsm の指紋生成＋既存指紋との差分レポート（データ行書込なし）

正本: docs/org/LV4_LANE_B_BULK_TEMPLATE_T0_APPROVAL.md
入口既定: 04…/09.SC純正バルクxlsm保存（人間がDLして保存→Agent確認後06へ）
出口既定: 04…/05.SC処理結果・ログ退避（人間）
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(2)

# 同一ディレクトリの指紋関数を再利用
from c1_packaged import compute_header_fingerprint  # noqa: E402

SCRIPT_DIR = Path(__file__).resolve().parent
LOG = logging.getLogger("c1_bulk_fingerprint")

DEFAULT_DRIVE04 = Path(r"G:/マイドライブ/04.amazonカタログ作成（CSV一括UL）")
DEFAULT_INBOX = DEFAULT_DRIVE04 / "09.SC純正バルクxlsm保存（人間がDLして保存→Agent確認後06へ）"
DEFAULT_REPORT = DEFAULT_DRIVE04 / "05.SC処理結果・ログ退避（人間）"
DEFAULT_FP_DIR = SCRIPT_DIR / "fingerprints"
DEFAULT_ROWS = [3, 4, 5]
SHEET_CANDIDATES = ("テンプレート", "Template", "template")


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _safe_name(name: str) -> str:
    s = re.sub(r"[^\w.\-]+", "_", name, flags=re.UNICODE)
    return s[:80] or "xlsm"


def _pick_sheet(wb: Any) -> Tuple[str, Any]:
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return name, wb[name]
    # フォールバック: 「テンプレ」を含む最初のシート
    for name in wb.sheetnames:
        if "テンプレ" in name or "template" in name.lower():
            return name, wb[name]
    raise RuntimeError("テンプレートシートが見つかりません: %s" % (wb.sheetnames,))


def _load_baselines(fp_dir: Path) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not fp_dir.is_dir():
        return out
    for path in sorted(fp_dir.glob("*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            LOG.warning("ベースライン読取失敗 %s: %s", path, e)
            continue
        if not data.get("sha256"):
            continue
        data["_baselinePath"] = str(path)
        data["_baselineName"] = path.name
        out.append(data)
    return out


def _resolve_inbox_file(inbox: Path) -> Path:
    if inbox.is_file():
        if inbox.suffix.lower() != ".xlsm":
            raise RuntimeError("xlsm ではありません: %s" % inbox)
        return inbox
    if not inbox.is_dir():
        raise RuntimeError("入口がありません: %s" % inbox)
    files = sorted(
        [p for p in inbox.iterdir() if p.is_file() and p.suffix.lower() == ".xlsm"],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not files:
        raise RuntimeError("入口に xlsm がありません: %s" % inbox)
    return files[0]


def fingerprint_one(
    xlsm_path: Path,
    baselines: List[Dict[str, Any]],
    rows: List[int],
    max_col_override: Optional[int],
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=False)
    try:
        sheet_name, ws = _pick_sheet(wb)
        max_col = int(max_col_override or ws.max_column or 300)
        sha = compute_header_fingerprint(ws, rows, max_col)
        comparisons: List[Dict[str, Any]] = []
        best: Optional[Dict[str, Any]] = None
        for b in baselines:
            b_sha = str(b.get("sha256") or "")
            b_rows = b.get("rows") or rows
            b_max = int(b.get("maxCol") or max_col)
            # ベースラインと同じ rows/maxCol でもう一度（比較公平）
            sha_cmp = compute_header_fingerprint(ws, list(b_rows), b_max)
            match = sha_cmp == b_sha
            entry = {
                "baselineName": b.get("_baselineName"),
                "baselinePath": b.get("_baselinePath"),
                "productType": b.get("productType"),
                "baselineSha256": b_sha,
                "comparedSha256": sha_cmp,
                "baselineRows": b_rows,
                "baselineMaxCol": b_max,
                "match": match,
            }
            comparisons.append(entry)
            if match:
                best = entry
        status = "match" if best else ("unknown_template" if baselines else "no_baseline")
        # データ開始行の粗い提案（空でない行を上から）
        data_start_guess = None
        for r in range(6, min(20, (ws.max_row or 20) + 1)):
            # 列1に何かあれば候補
            if ws.cell(r, 1).value not in (None, ""):
                # 注記っぽい長文はスキップ
                v = str(ws.cell(r, 1).value)
                if len(v) > 80:
                    continue
                data_start_guess = r
                break
        return {
            "status": status,
            "xlsmPath": str(xlsm_path),
            "xlsmName": xlsm_path.name,
            "sheetName": sheet_name,
            "sha256": sha,
            "rows": rows,
            "maxCol": max_col,
            "matchedBaseline": best,
            "comparisons": comparisons,
            "dataStartRowGuess": data_start_guess,
            "note": "B-T0: データ行・マスタへの書込なし。一致しなくてもレポート成果。",
        }
    finally:
        wb.close()


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="B-T0: 純正バルク xlsm 指紋＋既存指紋差分（書込なし）"
    )
    parser.add_argument(
        "--inbox",
        default=str(DEFAULT_INBOX),
        help="09フォルダまたは xlsm ファイルパス",
    )
    parser.add_argument(
        "--report-dir",
        default=str(DEFAULT_REPORT),
        help="05 レポート出力先",
    )
    parser.add_argument(
        "--fingerprints-dir",
        default=str(DEFAULT_FP_DIR),
        help="比較用 fingerprints/*.json",
    )
    parser.add_argument(
        "--rows",
        default="3,4,5",
        help="指紋行（カンマ区切り）",
    )
    parser.add_argument("--max-col", type=int, default=None, help="指紋列上限（省略時はシートmax）")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    rows = [int(x.strip()) for x in str(args.rows).split(",") if x.strip()]
    inbox = Path(args.inbox)
    report_dir = Path(args.report_dir)
    fp_dir = Path(args.fingerprints_dir)

    try:
        xlsm = _resolve_inbox_file(inbox)
    except RuntimeError as e:
        LOG.error("%s", e)
        return 2

    baselines = _load_baselines(fp_dir)
    if not baselines:
        LOG.error("ベースラインが0件です: %s", fp_dir)
        return 2

    LOG.info("対象 xlsm=%s", xlsm)
    LOG.info("ベースライン件数=%d dir=%s", len(baselines), fp_dir)

    try:
        result = fingerprint_one(xlsm, baselines, rows, args.max_col)
    except Exception as e:
        LOG.exception("指紋失敗: %s", e)
        report_dir.mkdir(parents=True, exist_ok=True)
        stamp = _utc_stamp()
        fail_path = report_dir / ("B_T0_%s_%s_FINGERPRINT.json" % (stamp, _safe_name(xlsm.stem)))
        fail_doc = {
            "runId": "B_T0_%s" % stamp,
            "status": "FAILED",
            "error": str(e),
            "xlsmPath": str(xlsm),
            "recordedAt": datetime.now(timezone.utc).isoformat(),
        }
        fail_path.write_text(json.dumps(fail_doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return 1

    stamp = _utc_stamp()
    run_id = "B_T0_%s" % stamp
    report_dir.mkdir(parents=True, exist_ok=True)
    safe = _safe_name(xlsm.stem)
    json_path = report_dir / ("B_T0_%s_%s_FINGERPRINT.json" % (stamp, safe))
    txt_path = report_dir / ("B_T0_%s_%s_SUMMARY.txt" % (stamp, safe))

    doc = {
        "runId": run_id,
        "stage": "B-T0",
        "recordedAt": datetime.now(timezone.utc).isoformat(),
        "inbox": str(inbox),
        "reportDir": str(report_dir),
        "fingerprintsDir": str(fp_dir),
        "result": result,
        "nextHumanStep": "問題なければ 09 のファイルを 06 へコピー／移動。データ行書込は B-T1。",
    }
    json_path.write_text(json.dumps(doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines = [
        "B-T0 fingerprint summary",
        "runId=%s" % run_id,
        "xlsm=%s" % xlsm.name,
        "status=%s" % result["status"],
        "sha256=%s" % result["sha256"],
        "sheet=%s" % result["sheetName"],
        "maxCol=%s" % result["maxCol"],
        "dataStartRowGuess=%s" % result.get("dataStartRowGuess"),
    ]
    if result.get("matchedBaseline"):
        mb = result["matchedBaseline"]
        lines.append(
            "match=%s productType=%s" % (mb.get("baselineName"), mb.get("productType"))
        )
    else:
        lines.append("match=none (unknown_template or mismatch)")
    for c in result.get("comparisons") or []:
        lines.append(
            "  - %s match=%s pt=%s" % (c.get("baselineName"), c.get("match"), c.get("productType"))
        )
    lines.append("json=%s" % json_path)
    lines.append(result.get("note") or "")
    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    LOG.info("status=%s sha256=%s", result["status"], result["sha256"])
    LOG.info("report json=%s", json_path)
    LOG.info("report summary=%s", txt_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
