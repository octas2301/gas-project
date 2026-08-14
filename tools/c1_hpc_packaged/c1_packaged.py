#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
C1 / C1-1b: 純正 xlsm → PACKAGED（HPC / FOOD系ローカル本線）

- GENERATED + マスタCSV併読（列名＝マスタ見出し）
- 必須マスタ欠落・URL空 → 親SKU一式除外
- 商品タックスコードはマスタ値（固定しない）
- 指紋不一致 → 本番停止／DRY_RUN警告
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(2)

SCRIPT_DIR = Path(__file__).resolve().parent
LOG = logging.getLogger("c1_packaged")


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def resolve_path(p: str, base: Path) -> Path:
    path = Path(p)
    if not path.is_absolute():
        path = (base / path).resolve()
    return path


def expand_sub_batch_id_(path_str: str, sub_batch_id: str) -> str:
    """config の {subBatchId} を実 ID に置換。プレースホルダ無しならそのまま。"""
    s = str(path_str or "")
    if "{subBatchId}" not in s:
        return s
    sid = str(sub_batch_id or "").strip()
    if not sid:
        raise ValueError(
            "generated_csv に {subBatchId} があります。"
            "config.sub_batch_id または --sub-batch を指定してください。"
        )
    return s.replace("{subBatchId}", sid)


def compute_header_fingerprint(ws, rows: List[int], max_col: int = 300) -> str:
    parts: List[str] = []
    for r in rows:
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            parts.append("%d:%d:%s" % (r, c, _cell_str(v)))
    raw = "\n".join(parts).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


_REQUIRED_GENERATED_HEADERS = (
    "parentSku",
    "sellerSku",
    "priceAmazon",
    "shippingTemplate",
    "variationRole",
)
_ROLE_TOKENS = frozenset({"parent", "child", "親", "子供", "子"})


def _normalize_list_price(raw: str) -> str:
    """税込み参考価格用。数字のみ残す。範囲・調査メモは無効。"""
    s = _cell_str(raw)
    if not s:
        return ""
    s = re.sub(r"[円￥¥,\s]", "", s)
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return s
    return ""


def load_generated_csv(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        raw_rows = list(csv.reader(f))
    if not raw_rows:
        raise ValueError("GENERATED CSV が空です: %s" % path)
    headers = [_cell_str(h) for h in raw_rows[0]]
    if not any(headers):
        raise ValueError("GENERATED CSV にヘッダーがありません: %s" % path)
    missing = [h for h in _REQUIRED_GENERATED_HEADERS if h not in headers]
    if missing:
        raise ValueError("GENERATED CSV 必須列不足: %s" % ", ".join(missing))
    n = len(headers)
    rows: List[Dict[str, str]] = []
    for i, raw in enumerate(raw_rows[1:], start=2):
        if not raw or not any(_cell_str(c) for c in raw):
            continue
        if len(raw) != n:
            raise ValueError(
                "GENERATED CSV 列数不一致: line=%d expected=%d actual=%d "
                "(カンマ不足・余分で shippingTemplate がずれます)" % (i, n, len(raw))
            )
        row = {headers[j]: _cell_str(raw[j] if j < len(raw) else "") for j in range(n)}
        ship = _cell_str(row.get("shippingTemplate"))
        role = _cell_str(row.get("variationRole")).lower()
        if ship.lower() in _ROLE_TOKENS or ship in _ROLE_TOKENS:
            raise ValueError(
                "GENERATED CSV shippingTemplate に親子ロールが入っています (line=%d value=%r)。"
                "列ずれの可能性。csv.writer で書き直してください。" % (i, ship)
            )
        if role and role not in ("parent", "child"):
            LOG.warning("GENERATED variationRole 非標準 line=%d value=%r", i, role)
        rows.append(row)
    return headers, rows


def _norm_header(name: Any) -> str:
    """Sheets書き出しで改行入り見出しになることがあるため正規化。"""
    return _cell_str(name).replace("\n", " ").replace("\r", " ")


def _pick_master_field(row: Dict[str, str], aliases: List[str]) -> str:
    for name in aliases:
        if name in row and _cell_str(row.get(name)):
            return _cell_str(row.get(name))
    # strip / 改行除去後の一致
    norm_map = {_norm_header(k): v for k, v in row.items()}
    for name in aliases:
        key = _norm_header(name)
        if key in norm_map and _cell_str(norm_map.get(key)):
            return _cell_str(norm_map.get(key))
    return ""


def _find_master_header_row(rows: List[List[str]], master_columns: dict) -> int:
    """Google Sheets全件CSVは先頭に注記行があり、列名行が1行目とは限らない。"""
    must = set()
    for key in ("parent_sku", "child_sku"):
        for a in master_columns.get(key) or []:
            must.add(_norm_header(a))
    if not must:
        must = {"親SKU", "子SKU"}
    for i, row in enumerate(rows[:40]):
        cells = {_norm_header(c) for c in row}
        if must.issubset(cells):
            return i
    raise ValueError(
        "マスタCSVに列名行（親SKU/子SKU）が見つかりません。Sheets書き出しか列名を確認してください。"
    )


def load_master_csv(path: Path, master_columns: dict) -> Dict[str, Dict[str, str]]:
    """子SKU優先・親SKUでも引ける index: sku -> normalized fields。"""
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        all_rows = list(csv.reader(f))
    if not all_rows:
        raise ValueError("マスタCSVが空です: %s" % path)
    header_idx = _find_master_header_row(all_rows, master_columns)
    raw_headers = all_rows[header_idx]
    # 重複列名は後ろを優先（マスタ後半の本番列が前半の旧列を上書き）
    headers: List[str] = []
    seen: Dict[str, int] = {}
    for h in raw_headers:
        key = _norm_header(h) or ("__empty_%d" % len(headers))
        if key in seen:
            headers[seen[key]] = "__dup_%d_%s" % (seen[key], key)
        seen[key] = len(headers)
        headers.append(key)

    index: Dict[str, Dict[str, str]] = {}
    data_rows = 0
    for raw in all_rows[header_idx + 1 :]:
        if not raw or not any(_cell_str(c) for c in raw):
            continue
        row = {
            headers[i]: _cell_str(raw[i] if i < len(raw) else "")
            for i in range(len(headers))
        }
        fields: Dict[str, str] = {}
        for key, aliases in master_columns.items():
            fields[key] = _pick_master_field(row, aliases)
        child = fields.get("child_sku") or ""
        parent = fields.get("parent_sku") or ""
        if not child and not parent:
            continue
        data_rows += 1
        if child:
            index[child] = fields
        if parent and parent not in index:
            index[parent] = fields
        if parent and not child:
            index[parent] = fields
    LOG.info(
        "マスタCSV読込: %s header_row=%d data_rows=%d rows_index=%d",
        path,
        header_idx + 1,
        data_rows,
        len(index),
    )
    return index


def group_by_parent(rows: List[Dict[str, str]]) -> Dict[str, Dict[str, Any]]:
    groups: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        parent = row.get("parentSku") or ""
        if not parent:
            continue
        g = groups.setdefault(parent, {"parent": None, "children": []})
        role = (row.get("variationRole") or "").lower()
        if role == "parent" or (not row.get("childSku") and row.get("sellerSku") == parent):
            g["parent"] = row
        else:
            g["children"].append(row)
    return groups


def sub_batch_id_from_generated(path: Path, rows: List[Dict[str, str]]) -> str:
    name = path.name
    m = re.match(r"^(.+)_GENERATED\.csv$", name, re.I)
    if m:
        return m.group(1)
    if rows and rows[0].get("track"):
        return "%s_%s" % (rows[0].get("track"), _utc_stamp())
    return "C1_%s" % _utc_stamp()


def _yes_no_jp(raw: str, default: str = "いいえ") -> str:
    s = _cell_str(raw).lower()
    if not s:
        return default
    if s in ("はい", "yes", "true", "1", "y"):
        return "はい"
    if s in ("いいえ", "no", "false", "0", "n"):
        return "いいえ"
    if "はい" in _cell_str(raw):
        return "はい"
    if "いいえ" in _cell_str(raw):
        return "いいえ"
    return default


def split_keywords(text: str, n: int = 5) -> List[str]:
    parts = re.split(r"[\s　,、/|]+", _cell_str(text))
    parts = [p for p in parts if p]
    if n <= 1:
        joined = " ".join(parts)
        return [joined] if joined else [""]
    out = parts[:n]
    while len(out) < n:
        out.append("")
    return out


def resolve_highlight(
    title: str,
    master: Dict[str, str],
    bullet1: str,
    colmap: dict,
) -> Tuple[str, str]:
    """ハイライト: タイトルが上限超なら空。優先は colmap.highlight_priority（既定: 楽天→Yahoo→箇条書き①）。"""
    max_chars = int(colmap.get("highlight_max_title_chars") or 75)
    if len(title) > max_chars:
        return "", "skipped_title_gt_%d" % max_chars
    priority = colmap.get("highlight_priority") or [
        "catch_rakuten",
        "catch_yahoo",
        "bullet1",
    ]
    for key in priority:
        if key == "bullet1":
            value = bullet1
        else:
            value = _cell_str(master.get(key))
        if value:
            return value, str(key)
    return "", "empty"


def extract_browse_node_id(browse_text: str) -> str:
    """表示文字列「… (71192051)」または素の数字 → Node ID（ルーティング／突合用）。"""
    s = _cell_str(browse_text)
    if not s:
        return ""
    m = re.search(r"\((\d{6,})\)\s*$", s)
    if m:
        return m.group(1)
    if s.isdigit() and len(s) >= 6:
        return s
    m = re.search(r"\b(\d{6,})\b", s)
    return m.group(1) if m else ""


_BROWSE_PATH_BY_NODE: Optional[Dict[str, str]] = None


def load_browse_path_by_node(colmap: Optional[dict] = None) -> Dict[str, str]:
    """データを閲覧する由来の BrowsePath（xlsmプルダウン表記）。"""
    global _BROWSE_PATH_BY_NODE
    if _BROWSE_PATH_BY_NODE is not None:
        return _BROWSE_PATH_BY_NODE
    rel = ""
    if colmap:
        rel = _cell_str(colmap.get("browse_catalog_path"))
    candidates = []
    if rel:
        candidates.append(Path(rel))
        candidates.append(SCRIPT_DIR / rel)
    candidates.append(SCRIPT_DIR / "shelf_browse_catalog.json")
    out: Dict[str, str] = {}
    for p in candidates:
        if not p.is_file():
            continue
        try:
            data = _load_json(p)
        except Exception:
            continue
        for row in data.get("rows") or []:
            nid = _cell_str(row.get("browseNodeId"))
            path = _cell_str(row.get("browsePath"))
            if nid and path:
                out[nid] = path
        if out:
            break
    _BROWSE_PATH_BY_NODE = out
    return out


def resolve_browse_xlsm_value(
    browse_raw: str,
    colmap: dict,
) -> Tuple[str, Optional[str]]:
    """xlsm の推奨ブラウズノード列用。純正プルダウンと同じ表記を返す。

    Dropdown Lists 実値は『BrowsePath (NodeId)』。
    例: 食品・飲料・お酒 > 缶詰・瓶詰 > 肉の缶詰・瓶詰 (71192051)
    短縮名や Node ID だけは不可。
    """
    raw = _cell_str(browse_raw)
    if not raw:
        return "", "browse empty"
    by_node = load_browse_path_by_node(colmap)
    node = extract_browse_node_id(raw)

    def dropdown_label(path: str, nid: str) -> str:
        p = _cell_str(path)
        n = _cell_str(nid)
        if not p:
            return ""
        if n and p.endswith("(%s)" % n):
            return p
        if n:
            return "%s (%s)" % (p, n)
        return p

    if node and node in by_node:
        return dropdown_label(by_node[node], node), None
    if ">" in raw:
        path_only = re.sub(r"\s*\(\d{6,}\)\s*$", "", raw).strip()
        nid2 = extract_browse_node_id(raw)
        if nid2 and nid2 in by_node:
            return dropdown_label(by_node[nid2], nid2), None
        for nid3, path in by_node.items():
            if path_only == path:
                return dropdown_label(path, nid3), None
        if nid2:
            return dropdown_label(path_only, nid2), None
        return path_only, None
    leaf = re.sub(r"\s*\(\d{6,}\)\s*$", "", raw).strip()
    matches = [
        (nid4, p) for nid4, p in by_node.items() if p.split(" > ")[-1] == leaf
    ]
    if len(matches) == 1:
        return dropdown_label(matches[0][1], matches[0][0]), None
    if len(matches) > 1:
        return "", "BrowsePath 複数候補（Node IDをマスタに含めてください）: %s" % leaf
    return "", (
        "推奨ブラウズノードをプルダウン表記(BrowsePath (NodeId))に解決できない: %r "
        "（棚 catalog / Dropdown Lists を確認）" % raw[:80]
    )


def resolve_product_type_browse(
    master: Dict[str, str],
    defaults: dict,
    colmap: dict,
) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """(pt, browse, error)。require 時はマスタ必須・既定埋込禁止。
    browse は xlsm プルダウンと同じ『BrowsePath (NodeId)』（数値IDや短縮名だけは不可）。
    """
    allow = set(
        colmap.get("p4b_c1_product_types")
        or ["SEASONING", "HEALTH_PERSONAL_CARE"]
    )
    pt_m = _cell_str(master.get("amazon_product_type"))
    browse_raw = _cell_str(master.get("amazon_browse_node"))
    require = bool(colmap.get("p4b_require_master_pt_browse"))
    if require:
        if not pt_m or not browse_raw:
            return None, None, (
                "Amazon Product Type / Browse Node 必須"
                "（マスタ空は不可・既定埋込禁止）"
            )
        browse_m, berr = resolve_browse_xlsm_value(browse_raw, colmap)
        if berr or not browse_m:
            return None, None, berr or "BrowsePath 解決失敗"
        if pt_m not in allow:
            return None, None, (
                "Amazon Product Type 本線非許可: %s（許可=%s）"
                % (pt_m, ",".join(sorted(allow)))
            )
        return pt_m, browse_m, None
    if pt_m and pt_m in allow:
        browse_m, _ = resolve_browse_xlsm_value(
            browse_raw or _cell_str(defaults.get("browse")), colmap
        )
        return pt_m, browse_m, None
    browse_m, _ = resolve_browse_xlsm_value(_cell_str(defaults.get("browse")), colmap)
    return (
        _cell_str(defaults.get("product_type")),
        browse_m,
        None,
    )


def master_for_sku(master_index: Dict[str, Dict[str, str]], sku: str, parent_sku: str) -> Dict[str, str]:
    """子行に空欄が多いSheetsマスタ向け: 親の値を継承し、子の非空だけ上書き。
    定価は親の非数字（調査メモ等）を子へ継承しない。
    """
    parent_fields = master_index.get(parent_sku) or {} if parent_sku else {}
    child_fields = master_index.get(sku) or {} if sku else {}
    if not parent_fields and not child_fields:
        return {}
    if sku and sku == parent_sku:
        return dict(parent_fields or child_fields)
    merged = dict(parent_fields)
    for k, v in child_fields.items():
        if _cell_str(v):
            merged[k] = v
    if not _cell_str(child_fields.get("list_price")):
        parent_lp = _cell_str(parent_fields.get("list_price"))
        merged["list_price"] = parent_lp if _normalize_list_price(parent_lp) else ""
    return merged


def resolve_url(
    row: Dict[str, str],
    url_override: Dict[str, str],
    master: Dict[str, str],
) -> str:
    """Amazon MAIN URL のみ。GENERATED / CDN へのフォールバック禁止（社長決定）。"""
    sku = row.get("sellerSku") or row.get("childSku") or row.get("parentSku") or ""
    if sku and sku in url_override:
        return _cell_str(url_override[sku])
    return _cell_str(master.get("amazon_main_url"))


def resolve_sub_urls(master: Dict[str, str], limit: int = 8) -> List[str]:
    """サブ画像は マスタ Amazon PT URL のみ。GENERATED / 楽天CDN へのフォールバック禁止。"""
    raw = _cell_str(master.get("amazon_pt_url"))
    if not raw:
        return []
    out: List[str] = []
    for part in re.split(r"[|\n,]+", raw):
        url = part.strip()
        if not url or not url.lower().startswith("http"):
            continue
        if url not in out:
            out.append(url)
        if len(out) >= limit:
            break
    return out


def resolve_size(row: Dict[str, str], size_map: Dict[str, str], master: Dict[str, str]) -> str:
    sku = row.get("sellerSku") or row.get("childSku") or ""
    if sku and sku in size_map:
        return _cell_str(size_map[sku])
    if _cell_str(master.get("variation_value")):
        return _cell_str(master.get("variation_value"))
    for key in ("variationValue", "size", "バリエーション値"):
        if _cell_str(row.get(key)):
            return _cell_str(row.get(key))
    return _cell_str(row.get("setCount"))


def parse_set_count(
    set_count_raw: str = "",
    size: str = "",
    gen_set_count: str = "",
) -> str:
    """A.セット商品数「3個で1セット」／サイズ「3缶/480g」／GENERATED setCount → 缶数。"""
    raw = _cell_str(set_count_raw)
    m = re.search(r"(\d+)\s*個", raw)
    if m:
        return m.group(1)
    m = re.search(r"(\d+)\s*缶", raw)
    if m:
        return m.group(1)
    size_s = _cell_str(size)
    m = re.search(r"(\d+)\s*缶", size_s)
    if m:
        return m.group(1)
    gen = _cell_str(gen_set_count)
    if gen.isdigit():
        return gen
    if raw.isdigit():
        return raw
    return ""


def parse_weight_from_size(size: str) -> str:
    """バリエーション値「3缶/480g」→ 480。"""
    s = _cell_str(size)
    m = re.search(r"/\s*(\d+(?:\.\d+)?)\s*g", s, re.I)
    if m:
        return m.group(1)
    m = re.search(r"(\d+(?:\.\d+)?)\s*g", s, re.I)
    if m:
        return m.group(1)
    return ""


def resolve_list_price(
    sku: str,
    parent_sku: str,
    master: Dict[str, str],
    list_price_override: Dict[str, str],
    master_index: Optional[Dict[str, Dict[str, str]]] = None,
    is_parent: bool = False,
    from_children: Optional[List[str]] = None,
    sale_price: str = "",
) -> str:
    """優先: override → 当SKU行の定価 →（親行なら）子の定価 → マージ定価 → 販売価格フォールバック。"""
    for key in (sku, parent_sku):
        if key and key in list_price_override:
            norm = _normalize_list_price(list_price_override[key])
            if norm:
                return norm
            LOG.warning("list_price_override が数字ではない sku=%s value=%r", key, list_price_override[key])

    if master_index and sku and sku in master_index:
        own = _normalize_list_price(master_index[sku].get("list_price") or "")
        if own:
            return own

    if is_parent:
        for lp in from_children or []:
            norm = _normalize_list_price(lp)
            if norm:
                return norm
        if master_index and parent_sku:
            for fields in master_index.values():
                if fields.get("parent_sku") != parent_sku or not fields.get("child_sku"):
                    continue
                norm = _normalize_list_price(fields.get("list_price") or "")
                if norm:
                    return norm

    from_master = _normalize_list_price(master.get("list_price") or "")
    if from_master:
        return from_master

    # 定価空のときのみ: GENERATED 販売価格 → マスタ販売価格amazon（master_columns key=price）
    for cand in (
        sale_price,
        master.get("price") or "",
        master.get("販売価格amazon") or "",
    ):
        fb = _normalize_list_price(cand)
        if fb:
            LOG.info(
                "list_price fallback to sale_price sku=%s value=%s",
                sku or parent_sku,
                fb,
            )
            return fb
    return ""


def resolve_shipping(gen_row: Dict[str, str], defaults: dict) -> str:
    ship = _cell_str(gen_row.get("shippingTemplate"))
    if not ship or ship.lower() in _ROLE_TOKENS or ship in _ROLE_TOKENS:
        fallback = defaults.get("shipping", "送料無料パターン")
        if ship:
            LOG.warning("shippingTemplate 不正 %r → 既定 %r", ship, fallback)
        return fallback
    return ship


def build_row_attrs(
    gen_row: Dict[str, str],
    master: Dict[str, str],
    defaults: dict,
    url_override: Dict[str, str],
    size_map: Dict[str, str],
    list_price_override: Dict[str, str],
    is_parent: bool,
    parent_sku: str,
    colmap: dict,
    master_index: Optional[Dict[str, Dict[str, str]]] = None,
    list_price_from_children: Optional[List[str]] = None,
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    sku = _cell_str(gen_row.get("sellerSku") or gen_row.get("childSku") or parent_sku)
    url = resolve_url(gen_row, url_override, master)
    if not url:
        return None, "Amazon MAIN URL 空（フォールバック禁止）"

    title = _cell_str(master.get("product_name_amazon"))
    if not title or "手入力" in title:
        title = _cell_str(gen_row.get("productName"))
    if not title or "手入力" in title:
        return None, "商品名不正/プレースホルダ"

    price = _cell_str(gen_row.get("priceAmazon"))
    if not price:
        return None, "priceAmazon 空"

    bullet1 = _cell_str(master.get("bullet1"))
    if not bullet1:
        return None, "マスタ 商品説明の箇条書き① 空"

    tax = _cell_str(master.get("tax_code"))
    if not tax:
        return None, "マスタ 商品タックスコード 空"

    origin = _cell_str(master.get("origin"))
    if not origin:
        return None, "マスタ 原産国 空"

    policy = colmap.get("c1_quantity_policy") or {}
    size = "" if is_parent else resolve_size(gen_row, size_map, master)
    if not is_parent and not size:
        return None, "サイズ/バリエーション値 空"

    set_count = ""
    if policy.get("parse_set_count"):
        set_count = parse_set_count(
            _cell_str(master.get("set_count")),
            size=size,
            gen_set_count=_cell_str(gen_row.get("setCount")),
        )
        if not is_parent and not set_count:
            return None, "セット数を解析できない（A.セット商品数／setCount／サイズ）"

    if policy.get("use_set_count_for_unit_count"):
        if is_parent:
            unit_count = ""
            unit_uom = ""
        else:
            unit_count = set_count
            unit_uom = (
                _cell_str(master.get("unit_uom"))
                or _cell_str(policy.get("unit_uom_default"))
                or _cell_str(defaults.get("unit_uom"))
            )
            if not unit_count or not unit_uom:
                return None, "ユニット数/単位 空（セット数ポリシー）"
    else:
        unit_count = _cell_str(master.get("unit_count")) or _cell_str(defaults.get("unit_count"))
        unit_uom = _cell_str(master.get("unit_uom")) or _cell_str(defaults.get("unit_uom"))
        if not unit_count or not unit_uom:
            return None, "マスタ ユニット数/単位 空"

    mfr_name = _cell_str(master.get("mfr_name"))
    if not mfr_name:
        return None, "マスタ メーカー名 空"

    mfr_part = (
        _cell_str(master.get("mfr_part"))
        or _cell_str(gen_row.get("manufacturerPart"))
        or ("" if is_parent else sku)
    )
    bullets = [
        bullet1,
        _cell_str(master.get("bullet2")) or bullet1,
        _cell_str(master.get("bullet3")) or bullet1,
        _cell_str(master.get("bullet4")) or bullet1,
        _cell_str(master.get("bullet5")) or bullet1,
    ]
    kw_slots = int(colmap.get("keyword_max_slots") or 5)
    kws = split_keywords(master.get("keywords") or "", n=kw_slots)

    heat = _yes_no_jp(master.get("heat") or "", defaults.get("heat", "いいえ"))
    liquid = _yes_no_jp(master.get("liquid") or "", defaults.get("liquid", "いいえ"))
    if policy.get("omit_color"):
        color = None
    else:
        color = _cell_str(master.get("color")) or defaults.get("color", "その他")
    list_price = resolve_list_price(
        sku,
        parent_sku,
        master,
        list_price_override,
        master_index=master_index,
        is_parent=is_parent,
        from_children=list_price_from_children,
        sale_price=price,
    )
    if not list_price:
        return None, (
            "税込み参考価格（定価）が数字でない。"
            "子SKU行の「定価、市場価格」を数字にするか、"
            "販売価格amazon／GENERATED priceAmazon を埋めてください"
            "（定価空時のみ販売価格へフォールバック）"
        )
    ingredients = _cell_str(master.get("ingredients"))
    passthrough: Dict[str, str] = {}
    for key in colmap.get("passthrough_fields") or []:
        if policy.get("omit_item_form") and key == "item_form":
            continue
        if policy.get("omit_temperature_rating") and key == "temperature_rating":
            continue
        if policy.get("temperature_fixed") and key == "temperature_rating":
            passthrough[key] = _cell_str(policy.get("temperature_fixed"))
            continue
        if policy.get("use_set_count_for_number_items") and key == "number_items":
            if not is_parent and set_count:
                passthrough[key] = set_count
            continue
        if policy.get("parse_weight_from_size") and key == "item_weight":
            if not is_parent:
                weight = parse_weight_from_size(size)
                if weight:
                    passthrough[key] = weight
                elif _cell_str(master.get("item_weight")):
                    # 子行に直接重量がある場合のみ（親総重量の継承は aliases 側で遮断）
                    passthrough[key] = _cell_str(master.get("item_weight"))
            continue
        if policy.get("parse_weight_from_size") and key == "item_weight_unit" and is_parent:
            continue
        if policy.get("omit_color") and key == "color":
            continue
        value = _cell_str(master.get(key)) or _cell_str(defaults.get(key))
        if value:
            passthrough[key] = value

    if (
        policy.get("parse_weight_from_size")
        and not is_parent
        and "item_weight" in (colmap.get("passthrough_fields") or [])
        and not passthrough.get("item_weight")
    ):
        return None, "商品の重量をサイズから解析できない"

    missing_extra = []
    skip_required = {
        "mfr_name", "bullet1", "tax_code", "origin",
        "unit_count", "unit_uom",
    }
    if policy.get("use_set_count_for_unit_count"):
        skip_required.add("unit_count")
    for key in colmap.get("required_master_fields") or []:
        if key in skip_required:
            continue
        value = _cell_str(master.get(key)) or _cell_str(defaults.get(key))
        if not value:
            missing_extra.append(key)
    if missing_extra:
        return None, "マスタ必須項目空: " + ",".join(missing_extra)

    pt_val, browse_val, pt_err = resolve_product_type_browse(master, defaults, colmap)
    if pt_err:
        return None, pt_err

    highlight, highlight_source = resolve_highlight(title, master, bullet1, colmap)
    LOG.info(
        "rowAttrs sku=%s pt=%s browseLen=%d highlightSrc=%s titleLen=%d",
        sku,
        pt_val or "",
        len(browse_val or ""),
        highlight_source,
        len(title),
    )

    return {
        "sku": sku,
        "url": url,
        "sub_urls": resolve_sub_urls(master),
        "title": title,
        "price": price,
        "size": size or None,
        "mfr": mfr_part or None,
        "mfr_name": mfr_name,
        "highlight": highlight,
        "highlight_source": highlight_source,
        "desc": bullet1,
        "specs": bullets,
        "keywords": kws,
        "color": color,
        "heat": heat,
        "liquid": liquid,
        "origin": origin,
        "tax_code": tax,
        "unit_count": unit_count,
        "unit_uom": unit_uom,
        "list_price": list_price,
        "ingredients": ingredients or None,
        "passthrough": passthrough,
        "shipping": resolve_shipping(gen_row, defaults),
        "inventory": _cell_str(gen_row.get("inventory") or "0"),
        "product_type": pt_val or "",
        "browse": browse_val or "",
        "amazon_product_type": _cell_str(master.get("amazon_product_type")),
        "amazon_browse_node": _cell_str(master.get("amazon_browse_node")),
    }, None


def evaluate_parent(
    parent_sku: str,
    bundle: Dict[str, Any],
    url_override: Dict[str, str],
    size_map: Dict[str, str],
    list_price_override: Dict[str, str],
    master_index: Dict[str, Dict[str, str]],
    defaults: dict,
    colmap: dict,
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    parent_row = bundle.get("parent")
    children = bundle.get("children") or []
    if not children:
        return None, "子SKUなし"
    if not parent_row:
        return None, "親行なし（GENERATED）"

    # 子を先に構築し、親の参考価格は子SKU行の定価を使う
    built_children = []
    for ch in children:
        sku = _cell_str(ch.get("sellerSku") or ch.get("childSku"))
        if not sku:
            return None, "子 sellerSku 空"
        ch_master = master_for_sku(master_index, sku, parent_sku)
        attrs, cerr = build_row_attrs(
            ch,
            ch_master,
            defaults,
            url_override,
            size_map,
            list_price_override,
            False,
            parent_sku,
            colmap,
            master_index=master_index,
        )
        if attrs is None:
            return None, "子(%s):%s" % (sku, cerr or "unknown")
        built_children.append(attrs)

    child_prices = [c.get("list_price") or "" for c in built_children]
    parent_master = master_for_sku(master_index, parent_sku, parent_sku)
    parent_attrs, err = build_row_attrs(
        parent_row,
        parent_master,
        defaults,
        url_override,
        size_map,
        list_price_override,
        True,
        parent_sku,
        colmap,
        master_index=master_index,
        list_price_from_children=child_prices,
    )
    if parent_attrs is None:
        return None, "親:" + (err or "unknown")

    for attrs in built_children:
        if not attrs["price"]:
            attrs["price"] = parent_attrs["price"]

    return {
        "parent_sku": parent_sku,
        "parent": parent_attrs,
        "children": built_children,
    }, None


def write_family_rows(ws, start_row: int, family: Dict[str, Any], colmap: dict, defaults: dict) -> int:
    cols = colmap["cols"]

    def setc(r: int, key: str, value: Any) -> None:
        if key not in cols:
            return
        if value is None or value == "":
            return
        ws.cell(r, cols[key]).value = value

    def write_common(r: int, attrs: Dict[str, Any], parentage: str, parent_sku_val: Any) -> None:
        pt_val = _cell_str(attrs.get("product_type")) or _cell_str(defaults.get("product_type"))
        browse_val = _cell_str(attrs.get("browse"))
        if not browse_val and not colmap.get("p4b_require_master_pt_browse"):
            browse_val = _cell_str(defaults.get("browse"))
        setc(r, "sku", attrs["sku"])
        setc(r, "product_type", pt_val)
        setc(r, "action", defaults["action"])
        setc(r, "parentage", parentage)
        setc(r, "parent_sku", parent_sku_val)
        setc(r, "var_theme", defaults["var_theme"])
        setc(r, "title", attrs["title"])
        setc(r, "highlight", attrs["highlight"])
        setc(r, "brand", defaults["brand"])
        setc(r, "id_type", defaults["id_type"])
        setc(r, "browse", browse_val)
        setc(r, "mfr_name", attrs["mfr_name"])
        setc(r, "main_image_url", attrs["url"])
        for i, sub_url in enumerate(attrs.get("sub_urls") or []):
            setc(r, "other_image%d" % (i + 1), sub_url)
        setc(r, "desc", attrs["desc"])
        specs = attrs.get("specs") or []
        for i, key in enumerate(["spec1", "spec2", "spec3", "spec4", "spec5"]):
            if i < len(specs):
                setc(r, key, specs[i])
        kws = attrs.get("keywords") or []
        for i, key in enumerate(["keyword1", "keyword2", "keyword3", "keyword4", "keyword5"]):
            if i < len(kws) and kws[i]:
                setc(r, key, kws[i])
        setc(r, "color", attrs.get("color"))
        setc(r, "size", attrs.get("size"))
        setc(r, "mfr_part", attrs.get("mfr"))
        setc(r, "import_type", defaults.get("import_type"))
        setc(r, "exclusive", defaults.get("exclusive"))
        setc(r, "heat", attrs.get("heat") or defaults.get("heat"))
        setc(r, "ingredients", attrs.get("ingredients"))
        setc(r, "unit_count", attrs.get("unit_count"))
        setc(r, "unit_uom", attrs.get("unit_uom"))
        setc(r, "condition", defaults.get("condition"))
        setc(r, "list_price", attrs.get("list_price"))
        setc(r, "tax_code", attrs.get("tax_code"))
        setc(r, "fulfillment", defaults.get("fulfillment"))
        inv = attrs.get("inventory") or "0"
        setc(r, "inventory", int(inv) if str(inv).isdigit() else inv)
        price = attrs["price"]
        setc(r, "price", float(price) if _is_number(price) else price)
        setc(r, "shipping", attrs.get("shipping") or defaults.get("shipping"))
        setc(r, "origin", attrs.get("origin"))
        setc(r, "battery_needed", defaults.get("battery_needed"))
        setc(r, "battery_included", defaults.get("battery_included"))
        setc(r, "hazmat", defaults.get("hazmat"))
        setc(r, "liquid", attrs.get("liquid") or defaults.get("liquid"))
        for key, value in (attrs.get("passthrough") or {}).items():
            setc(r, key, value)

    r = start_row
    write_common(r, family["parent"], "親", None)
    r += 1
    for ch in family["children"]:
        write_common(r, ch, "子供", family["parent_sku"])
        r += 1
    return r


def _is_number(s: str) -> bool:
    try:
        float(s)
        return True
    except (TypeError, ValueError):
        return False


def clear_data_rows(ws, start_row: int, max_col: Optional[int] = None) -> None:
    if ws.max_row < start_row:
        return
    clear_to = int(max_col or ws.max_column)
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, clear_to + 1):
            ws.cell(r, c).value = None


def build_mapping_report(families: List[Dict[str, Any]], colmap: dict) -> List[dict]:
    cols = colmap["cols"]
    report = []
    for fam in families:
        p = fam["parent"]
        report.append(
            {
                "parentSku": fam["parent_sku"],
                "childCount": len(fam["children"]),
                "taxCode": p.get("tax_code"),
                "mfrName": p.get("mfr_name"),
                "mainImageCol": cols.get("main_image_url"),
                "subImageCount": len(p.get("sub_urls") or []),
                "children": [
                    {
                        "sku": c["sku"],
                        "size": c.get("size"),
                        "urlPresent": bool(c.get("url")),
                        "subImageCount": len(c.get("sub_urls") or []),
                        "taxCode": c.get("tax_code"),
                    }
                    for c in fam["children"]
                ],
            }
        )
    return report


def run(config_path: Path, mode: str, sub_batch_id: Optional[str] = None) -> int:
    cfg = _load_json(config_path)
    base = config_path.parent

    mode = (mode or cfg.get("mode") or "dry_run").strip().lower()
    if mode not in ("dry_run", "prod"):
        raise SystemExit("mode は dry_run または prod")

    sid = str(sub_batch_id or cfg.get("sub_batch_id") or "").strip()
    template_path = resolve_path(cfg["template_path"], base)
    generated_csv = resolve_path(expand_sub_batch_id_(cfg["generated_csv"], sid), base)
    output_dir = resolve_path(cfg["output_dir"], base)
    log_dir = resolve_path(cfg.get("log_dir") or str(output_dir), base)
    fp_path = resolve_path(cfg.get("fingerprint_path") or "fingerprints/hpc_header_r3_r5.json", base)
    cm_raw = cfg.get("column_map_path") or "hpc_column_map.json"
    map_path = Path(cm_raw)
    if not map_path.is_absolute():
        cand = (base / map_path).resolve()
        map_path = cand if cand.is_file() else (SCRIPT_DIR / map_path).resolve()

    colmap = _load_json(map_path)
    defaults = dict(colmap.get("defaults") or {})
    defaults.update(cfg.get("defaults") or {})
    master_columns = colmap.get("master_columns") or {}

    master_index: Dict[str, Dict[str, str]] = {}
    master_csv_path = cfg.get("master_csv") or ""
    if master_csv_path:
        mp = resolve_path(master_csv_path, base)
        if not mp.is_file():
            raise FileNotFoundError("マスタCSVがありません: %s" % mp)
        master_index = load_master_csv(mp, master_columns)
    else:
        LOG.warning("master_csv 未設定 — C1-1b必須列は欠落し親除外の可能性大")

    url_override = {str(k): str(v) for k, v in (cfg.get("url_override_map") or {}).items()}
    size_map = {str(k): str(v) for k, v in (cfg.get("size_map") or {}).items()}
    list_price_override = {
        str(k): str(v) for k, v in (cfg.get("list_price_override_map") or {}).items()
    }
    parent_filter = set(cfg.get("parent_sku_filter") or [])

    output_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)

    if not template_path.is_file():
        raise FileNotFoundError("テンプレがありません: %s" % template_path)
    if not generated_csv.is_file():
        raise FileNotFoundError("GENERATED CSVがありません: %s" % generated_csv)

    headers, gen_rows = load_generated_csv(generated_csv)
    sub_batch_id = cfg.get("sub_batch_id") or sub_batch_id_from_generated(generated_csv, gen_rows)
    run_id = "C1_%s_%s" % (sub_batch_id, _utc_stamp())

    groups = group_by_parent(gen_rows)
    if parent_filter:
        groups = {k: v for k, v in groups.items() if k in parent_filter}

    accepted: List[Dict[str, Any]] = []
    excluded: List[Dict[str, str]] = []
    for parent_sku, bundle in sorted(groups.items()):
        fam, reason = evaluate_parent(
            parent_sku,
            bundle,
            url_override,
            size_map,
            list_price_override,
            master_index,
            defaults,
            colmap,
        )
        if fam is None:
            excluded.append({"parentSku": parent_sku, "reason": reason or "unknown"})
            LOG.warning("EXCLUDE parent=%s reason=%s", parent_sku, reason)
        else:
            accepted.append(fam)

    wb_probe = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
    sheet_name = colmap.get("sheet_name") or "テンプレート"
    if sheet_name not in wb_probe.sheetnames:
        raise RuntimeError("シートがありません: %s / %s" % (sheet_name, wb_probe.sheetnames))
    ws_probe = wb_probe[sheet_name]
    fp_rows = colmap.get("header_fingerprint_rows") or [3, 4, 5]
    fingerprint_max_col = int(colmap.get("fingerprint_max_col") or ws_probe.max_column)
    current_fp = compute_header_fingerprint(ws_probe, fp_rows, fingerprint_max_col)
    wb_probe.close()

    fp_status = "missing_baseline"
    fp_ok = True
    if fp_path.is_file():
        baseline = _load_json(fp_path)
        expected = baseline.get("sha256") or ""
        if expected and expected != current_fp:
            fp_ok = False
            fp_status = "mismatch"
            LOG.error("テンプレ指紋不一致 expected=%s current=%s", expected, current_fp)
        else:
            fp_status = "match"
            LOG.info("テンプレ指紋一致 sha256=%s", current_fp)
    else:
        LOG.warning("指紋ベースライン無し: %s", fp_path)

    report = {
        "runId": run_id,
        "mode": mode,
        "version": "C1-1c",
        "profile": colmap.get("profile") or defaults.get("product_type") or "HPC",
        "subBatchId": sub_batch_id,
        "templatePath": str(template_path),
        "generatedCsv": str(generated_csv),
        "masterCsv": master_csv_path or None,
        "generatedHeaders": headers,
        "fingerprint": {
            "status": fp_status,
            "sha256": current_fp,
            "baselinePath": str(fp_path),
            "rows": fp_rows,
            "maxCol": fingerprint_max_col,
        },
        "acceptedParents": [f["parent_sku"] for f in accepted],
        "excludedParents": excluded,
        "mapping": build_mapping_report(accepted, colmap),
        "counts": {
            "acceptedParents": len(accepted),
            "excludedParents": len(excluded),
            "acceptedChildren": sum(len(f["children"]) for f in accepted),
        },
        "sanctuary": {
            "masterWrite": False,
            "templateOverwrite": False,
            "outputOverwrite": False,
            "urlFallback": False,
            "taxCodeFixed": False,
        },
    }

    report_path = log_dir / ("%s_C1_REPORT.json" % run_id)
    _save_json(report_path, report)
    LOG.info("レポート: %s", report_path)

    if mode == "prod" and not fp_ok:
        LOG.error("本番停止: テンプレ指紋不一致")
        return 3
    if mode == "prod" and fp_status == "missing_baseline":
        LOG.error("本番停止: 指紋ベースラインなし")
        return 3
    if mode == "prod" and not accepted:
        LOG.error("本番停止: 出力対象の親が0件")
        return 4

    write_xlsm = mode == "prod" or bool(cfg.get("write_dryrun_xlsm"))
    if not write_xlsm:
        LOG.info("DRY_RUN: xlsm 非作成")
        return 0 if accepted or excluded else 1

    if mode == "dry_run" and not fp_ok:
        LOG.warning("DRY_RUN: 指紋不一致だが継続")

    if not accepted:
        LOG.warning("accepted=0 — xlsm 非作成")
        return 1

    suffix = "_DRYRUN" if mode == "dry_run" else ""
    parent_tag = accepted[0]["parent_sku"] if len(accepted) == 1 else ("n%d" % len(accepted))
    output_label = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        str(colmap.get("output_label") or defaults.get("product_type") or "HPC"),
    ).strip("_") or "C1"
    out_name = "%s_PACKAGED_%s_%s%s.xlsm" % (
        sub_batch_id, output_label, parent_tag, suffix
    )
    out_path = output_dir / out_name
    if out_path.exists():
        out_path = output_dir / (
            "%s_PACKAGED_%s_%s_%s%s.xlsm" %
            (sub_batch_id, output_label, parent_tag, _utc_stamp(), suffix)
        )

    work = output_dir / ("_work_%s.xlsm" % run_id)
    shutil.copy2(template_path, work)
    LOG.info("テンプレコピー: %s -> %s", template_path, work)

    try:
        wb = openpyxl.load_workbook(work, keep_vba=True, data_only=False)
        ws = wb[sheet_name]
        data_start = int(colmap.get("data_start_row") or 7)
        sample_row = int(colmap.get("sample_row") or 6)
        clear_data_rows(ws, data_start, int(colmap.get("clear_max_col") or ws.max_column))
        row_cursor = data_start
        for fam in accepted:
            row_cursor = write_family_rows(ws, row_cursor, fam, colmap, defaults)
        wb.save(work)
        wb.close()
        shutil.move(str(work), str(out_path))
        LOG.info("出力: %s", out_path)
        report["outputPath"] = str(out_path)
        report["rowsWritten"] = row_cursor - data_start
        report["sampleRowPreserved"] = sample_row
        _save_json(report_path, report)
    except Exception:
        if work.exists():
            fail_keep = log_dir / ("%s_PARTIAL.xlsm" % run_id)
            try:
                shutil.move(str(work), str(fail_keep))
                LOG.error("途中失敗: 部分ファイルを残しました %s", fail_keep)
                report["partialPath"] = str(fail_keep)
                _save_json(report_path, report)
            except Exception:
                LOG.exception("部分ファイルの退避にも失敗")
        raise

    return 0


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="C1 PACKAGED builder (HPC / FOOD系)")
    parser.add_argument("--config", required=True, help="config.json パス")
    parser.add_argument("--mode", choices=["dry_run", "prod"], default=None)
    parser.add_argument(
        "--sub-batch",
        default=None,
        help="generated_csv の {subBatchId} 置換（config.sub_batch_id より優先）",
    )
    parser.add_argument("--write-fingerprint", action="store_true")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    config_path = Path(args.config).resolve()
    cfg = _load_json(config_path)
    base = config_path.parent

    if args.write_fingerprint:
        template_path = resolve_path(cfg["template_path"], base)
        map_path = resolve_path(cfg.get("column_map_path") or str(SCRIPT_DIR / "hpc_column_map.json"), base)
        if not map_path.exists():
            map_path = SCRIPT_DIR / "hpc_column_map.json"
        colmap = _load_json(map_path)
        fp_path = resolve_path(cfg.get("fingerprint_path") or "fingerprints/hpc_header_r3_r5.json", base)
        wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
        ws = wb[colmap.get("sheet_name") or "テンプレート"]
        rows = colmap.get("header_fingerprint_rows") or [3, 4, 5]
        max_col = int(colmap.get("fingerprint_max_col") or ws.max_column)
        sha = compute_header_fingerprint(ws, rows, max_col)
        wb.close()
        _save_json(
            fp_path,
            {
                "sha256": sha,
                "rows": rows,
                "templatePath": str(template_path),
                "recordedAt": datetime.now(timezone.utc).isoformat(),
                "maxCol": max_col,
                "productType": (
                    colmap.get("profile") or
                    (colmap.get("defaults") or {}).get("product_type") or
                    "HEALTH_PERSONAL_CARE"
                ),
            },
        )
        LOG.info("指紋を保存: %s sha256=%s", fp_path, sha)
        return 0

    return run(
        config_path,
        args.mode or cfg.get("mode") or "dry_run",
        sub_batch_id=args.sub_batch,
    )


if __name__ == "__main__":
    sys.exit(main())
