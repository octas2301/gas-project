#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B-T1: 棚引き → 項目名で列解決 → dry_run／prod 充填（既存 c1_packaged.run を利用）。

正本:
  dry_run … docs/org/LV4_LANE_B_BULK_TEMPLATE_T1_APPROVAL.md
  prod    … docs/org/LV4_LANE_B_BULK_TEMPLATE_T1_PROD_APPROVAL.md
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(2)

from c1_bulk_name_map import resolve_cols_by_aliases  # noqa: E402
from c1_bulk_shelf_lookup import (  # noqa: E402
    DEFAULT_REPORT,
    DEFAULT_SHELF,
    SCRIPT_DIR,
    lookup_shelf,
)
from c1_packaged import run as c1_run  # noqa: E402

LOG = logging.getLogger("c1_bulk_fill_by_name")


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _pick_sheet(wb: Any, sheet_name: str) -> Any:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    for name in ("テンプレート", "Template", "template"):
        if name in wb.sheetnames:
            return wb[name]
    raise RuntimeError("シートなし: %s" % (wb.sheetnames,))


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="B-T1 name-map fill (dry_run / prod)")
    parser.add_argument(
        "--mode",
        choices=["dry_run", "prod"],
        default="dry_run",
        help="既定 dry_run。prod は棚 DL_NOT_NEEDED＋指紋一致必須",
    )
    parser.add_argument("--product-type", default="SEASONING")
    parser.add_argument("--shelf-dir", default=str(DEFAULT_SHELF))
    parser.add_argument("--registry", default=str(SCRIPT_DIR / "shelf_registry.json"))
    parser.add_argument("--report-dir", default=str(DEFAULT_REPORT))
    parser.add_argument(
        "--template",
        default="",
        help="棚引きをスキップしてテンプレを直接指定（dry_run 検証用）",
    )
    parser.add_argument("--generated-csv", required=True)
    parser.add_argument("--master-csv", default="")
    parser.add_argument("--output-dir", required=True, help="03 等")
    parser.add_argument("--column-map", default="")
    parser.add_argument("--fingerprint-path", default="")
    parser.add_argument("--parent-sku-filter", nargs="*", default=[])
    parser.add_argument(
        "--allow-dl-required",
        action="store_true",
        help="dry_run のみ: 棚なしでも --template 指定時に続行。prod では拒否",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    mode = (args.mode or "dry_run").strip().lower()
    if mode not in ("dry_run", "prod"):
        raise SystemExit("mode は dry_run または prod")

    if mode == "prod" and args.allow_dl_required:
        LOG.error("prod では --allow-dl-required は使えません（棚登録後に再実行）")
        return 2

    report_dir = Path(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    run_id = "B_T1_FILL_%s_%s_%s" % (mode.upper(), args.product_type, _utc_stamp())

    registry = _load_json(Path(args.registry))
    shelf = lookup_shelf(args.product_type, Path(args.shelf_dir), registry)
    shelf["runId"] = run_id + "_SHELF"
    _save_json(report_dir / ("%s_SHELF.json" % run_id), shelf)

    template_path = Path(args.template) if args.template else None

    if mode == "prod":
        # prod: 棚一致必須。手動 --template も指紋は c1_packaged 側で一致必須。
        if shelf.get("status") != "DL_NOT_NEEDED":
            LOG.error(
                "prod 中止: 棚が DL_NOT_NEEDED ではありません status=%s message=%s",
                shelf.get("status"),
                shelf.get("message"),
            )
            summary = report_dir / ("%s_SUMMARY.txt" % run_id)
            summary.write_text(
                "status=ABORTED_PROD_SHELF\nshelfStatus=%s\nmessage=%s\n"
                % (shelf.get("status"), shelf.get("message")),
                encoding="utf-8",
            )
            return 2
        if not template_path:
            template_path = Path(str(shelf["templatePath"]))
        LOG.info("prod 棚引き: DL不要 → %s", template_path)
    elif shelf.get("status") == "DL_NOT_NEEDED" and not template_path:
        template_path = Path(str(shelf["templatePath"]))
        LOG.info("棚引き: DL不要 → %s", template_path)
    elif shelf.get("status") == "DL_REQUIRED" and not template_path:
        LOG.error("%s", shelf.get("message"))
        summary = report_dir / ("%s_SUMMARY.txt" % run_id)
        summary.write_text(
            "status=ABORTED_DL_REQUIRED\nmessage=%s\n" % shelf.get("message"),
            encoding="utf-8",
        )
        return 2
    elif template_path and shelf.get("status") == "DL_REQUIRED" and not args.allow_dl_required:
        LOG.error("棚なし。--allow-dl-required と --template が必要です。")
        return 2

    assert template_path is not None
    if not template_path.is_file():
        raise FileNotFoundError(template_path)

    entry = shelf.get("registryEntry") or {}
    cm_rel = args.column_map or entry.get("columnMapPath") or "food_seasoning_column_map.json"
    cm_path = Path(cm_rel)
    if not cm_path.is_file():
        cm_path = SCRIPT_DIR / cm_rel
    colmap = _load_json(cm_path)
    aliases = colmap.get("xlsm_header_aliases") or {}
    if not aliases:
        raise SystemExit("column_map に xlsm_header_aliases がありません: %s" % cm_path)

    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb, str(colmap.get("sheet_name") or "テンプレート"))
        max_col = int(colmap.get("fingerprint_max_col") or colmap.get("clear_max_col") or ws.max_column)
        resolved, gaps = resolve_cols_by_aliases(
            ws,
            aliases,
            header_rows=list(colmap.get("header_alias_rows") or [4, 5]),
            max_col=max_col,
            legacy_cols=colmap.get("cols") or {},
        )
    finally:
        wb.close()

    miss_required = [g for g in gaps if g.get("required") and g.get("status") == "MISS"]
    gap_report = {
        "runId": run_id,
        "mode": mode,
        "templatePath": str(template_path),
        "columnMapPath": str(cm_path),
        "resolvedCount": len(resolved),
        "aliasCount": len(aliases),
        "missRequired": miss_required,
        "gaps": gaps,
        "shelf": {
            "status": shelf.get("status"),
            "reason": shelf.get("reason"),
            "message": shelf.get("message"),
        },
    }
    gap_path = report_dir / ("%s_NAME_MAP_GAPS.json" % run_id)
    _save_json(gap_path, gap_report)
    LOG.info(
        "項目名解決: resolved=%d/%d missRequired=%d → %s",
        len(resolved),
        len(aliases),
        len(miss_required),
        gap_path,
    )
    if miss_required:
        LOG.error("必須列が項目名で解決できません。充填中止。")
        return 3

    work_colmap = dict(colmap)
    work_colmap["cols"] = resolved
    work_colmap["cols_legacy_reference"] = colmap.get("cols") or {}

    fp_rel = (
        args.fingerprint_path
        or entry.get("fingerprintBaselinePath")
        or "fingerprints/food_seasoning_header_r3_r5.json"
    )
    fp_path = Path(fp_rel)
    if not fp_path.is_file():
        fp_path = SCRIPT_DIR / fp_rel

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="b_t1_") as tmp:
        tmp_dir = Path(tmp)
        work_map_path = tmp_dir / "column_map_resolved.json"
        _save_json(work_map_path, work_colmap)

        cfg = {
            "template_path": str(template_path),
            "generated_csv": str(Path(args.generated_csv).resolve()),
            "master_csv": str(Path(args.master_csv).resolve()) if args.master_csv else "",
            "output_dir": str(output_dir.resolve()),
            "log_dir": str(report_dir.resolve()),
            "fingerprint_path": str(fp_path.resolve()),
            "column_map_path": str(work_map_path),
            "mode": mode,
            "write_dryrun_xlsm": mode == "dry_run",
            "parent_sku_filter": list(args.parent_sku_filter or []),
            "size_map": {},
            "url_override_map": {},
            "b_t1_run_id": run_id,
        }
        cfg_path = tmp_dir / "config.json"
        _save_json(cfg_path, cfg)
        rc = c1_run(cfg_path, mode, sub_batch_id=None)

    summary = report_dir / ("%s_SUMMARY.txt" % run_id)
    summary.write_text(
        "runId=%s\nmode=%s\nshelfStatus=%s\ntemplate=%s\nresolved=%d\nmissRequired=%d\nc1_rc=%s\n"
        % (
            run_id,
            mode,
            shelf.get("status"),
            template_path,
            len(resolved),
            len(miss_required),
            rc,
        ),
        encoding="utf-8",
    )
    LOG.info("SUMMARY: %s", summary)
    return int(rc)


if __name__ == "__main__":
    sys.exit(main())
