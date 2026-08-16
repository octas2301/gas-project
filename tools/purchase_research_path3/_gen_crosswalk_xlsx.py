# -*- coding: utf-8 -*-
"""mall_keepa_field_crosswalk.xlsx を色・タグ・公式全項目付きで再生成。"""
from pathlib import Path
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.series import SeriesLabel
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import TableColumn

OUT = Path(__file__).resolve().parent / "mall_keepa_field_crosswalk.xlsx"
CSV_IN = Path(__file__).resolve().parent / "mall_keepa_field_crosswalk.csv"
DOCS_DIR = Path(__file__).resolve().parents[2] / "docs" / "org" / "competitor_fields"

GREEN = PatternFill("solid", fgColor="C6EFCE")  # ◎ P0
YELLOW = PatternFill("solid", fgColor="FFEB9C")  # ○ P1
ORANGE = PatternFill("solid", fgColor="FCE4D6")  # △ P2
GRAY = PatternFill("solid", fgColor="D9D9D9")  # × P3
BLUE = PatternFill("solid", fgColor="DDEBF7")  # Keepa専用
HEADER = PatternFill("solid", fgColor="305496")
HFONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

PICK_FILL = {"◎": GREEN, "○": YELLOW, "△": ORANGE, "×": GRAY, "K": BLUE}


def fill_pick(v):
    if v.startswith("◎"):
        return GREEN
    if v.startswith("○"):
        return YELLOW
    if v.startswith("△"):
        return ORANGE
    if v.startswith("×"):
        return GRAY
    if v.startswith("K"):
        return BLUE
    return None


# existing logical rows: append tags
# agent_pick, priority, effect_setcount, effect_price, effect_human, flow_A, flow_B2, flow_B76, flow_CPO, flow_FBA, flow_filter, flow_tags
TAGS = {
    "JAN/EAN": ("◎使えそう", "P0", "高", "中", "高", "Y", "Y", "Y", "Y", "N", "N", "A_Keepa,B_Step2,B_Step76,CPO"),
    "ASIN": ("K Keepa専用", "P0", "中", "中", "中", "Y", "N", "N", "N", "Y", "N", "A_Keepa,FBA"),
    "親ASIN/バリエーション": ("○補助", "P1", "高", "低", "中", "Y", "N", "N", "N", "N", "N", "A_Keepa"),
    "モール商品コード": ("○補助", "P1", "低", "低", "低", "Y", "Y", "N", "N", "N", "N", "A_Keepa,B_Step2"),
    "商品名": ("◎使えそう", "P0", "高", "中", "高", "Y", "Y", "N", "Y", "N", "Y", "A_Keepa,B_Step2,CPO,除外フィルタ"),
    "キャッチ/説明短文": ("○補助", "P1", "中", "低", "中", "Y", "Y", "N", "N", "N", "Y", "A_Keepa,B_Step2"),
    "メーカー": ("○補助", "P1", "低", "低", "中", "Y", "N", "N", "N", "N", "N", "A_Keepa"),
    "ブランド": ("△後続", "P2", "低", "低", "低", "N", "N", "Y", "N", "N", "N", "B_Step76"),
    "表示価格": ("◎使えそう", "P0", "高", "高", "高", "Y", "Y", "N", "Y", "N", "N", "A_Keepa,B_Step2,CPO"),
    "プレミアム/会員価格": ("△後続", "P2", "低", "中", "低", "N", "Y", "N", "Y", "N", "N", "B_Step2,CPO"),
    "参考価格/定価": ("△後続", "P2", "低", "低", "低", "Y", "N", "N", "N", "N", "Y", "A_Keepa,除外フィルタ"),
    "ポイント還元": ("◎使えそう", "P0", "中", "高", "中", "Y", "Y", "N", "Y", "N", "N", "A_Keepa,B_Step2,CPO"),
    "送料フラグ": ("◎使えそう", "P0", "中", "高", "中", "Y", "Y", "N", "Y", "N", "Y", "A_Keepa,B_Step2,CPO,除外フィルタ"),
    "実質価格（派生）": ("◎使えそう", "P0", "高", "高", "高", "Y", "Y", "N", "Y", "N", "N", "A_Keepa,B_Step2,CPO"),
    "商品URL": ("○補助", "P1", "低", "低", "高", "Y", "Y", "N", "Y", "N", "N", "A_Keepa,B_Step2,CPO"),
    "画像URL": ("◎使えそう", "P0", "高", "低", "高", "Y", "Y", "N", "N", "N", "N", "A_Keepa,B_Step2"),
    "販売店": ("◎使えそう", "P0", "中", "低", "中", "Y", "Y", "N", "N", "N", "Y", "A_Keepa,B_Step2,除外フィルタ"),
    "在庫有無": ("○補助", "P1", "中", "低", "低", "Y", "Y", "N", "N", "N", "Y", "A_Keepa,B_Step2,除外フィルタ"),
    "新品/中古": ("○補助", "P1", "中", "低", "低", "Y", "Y", "N", "N", "N", "Y", "A_Keepa,B_Step2,除外フィルタ"),
    "レビュー件数": ("○補助", "P1", "中", "低", "中", "Y", "Y", "Y", "N", "N", "N", "A_Keepa,B_Step2,B_Step76"),
    "レビュー点": ("△後続", "P2", "低", "低", "低", "N", "Y", "N", "N", "N", "Y", "B_Step2,除外フィルタ"),
    "カテゴリID": ("△後続", "P2", "低", "低", "低", "N", "N", "Y", "N", "N", "N", "B_Step76"),
    "カテゴリ名パス": ("△後続", "P2", "低", "低", "低", "N", "N", "Y", "N", "N", "N", "B_Step76"),
    "ブランドID": ("△後続", "P2", "低", "低", "低", "N", "N", "Y", "N", "N", "N", "B_Step76"),
    "発売日": ("△後続", "P2", "低", "低", "低", "Y", "N", "N", "N", "N", "N", "A_Keepa"),
    "梱包寸法重量": ("K Keepa専用", "P0", "無", "無", "中", "Y", "N", "N", "N", "Y", "N", "A_Keepa,FBA"),
    "内容個数API": ("○補助", "P1", "高", "低", "高", "Y", "N", "N", "N", "N", "N", "A_Keepa"),
    "売れ筋": ("K Keepa専用", "P1", "無", "無", "中", "Y", "N", "N", "N", "N", "N", "A_Keepa"),
    "FBA手数料": ("K Keepa専用", "P1", "無", "無", "低", "N", "N", "N", "N", "Y", "N", "FBA"),
    "検索メタ": ("○補助", "P1", "低", "低", "低", "Y", "Y", "N", "N", "N", "N", "A_Keepa,B_Step2"),
    "アダルト": ("○補助", "P1", "低", "低", "低", "N", "Y", "N", "N", "N", "Y", "B_Step2,除外フィルタ"),
    "ふるさと納税": ("◎使えそう", "P0", "高", "高", "高", "Y", "Y", "N", "Y", "N", "Y", "A_Keepa,B_Step2,CPO,除外フィルタ"),
    "選択式バリエーション": ("◎使えそう", "P0", "高", "高", "高", "Y", "Y", "N", "Y", "N", "Y", "A_Keepa,B_Step2,CPO,除外フィルタ"),
    "ポイント期間": ("○補助", "P1", "低", "高", "低", "Y", "Y", "N", "Y", "N", "N", "A_Keepa,B_Step2,CPO"),
    "あす楽": ("×今回使わない", "P3", "無", "無", "無", "N", "N", "N", "N", "N", "N", ""),
    "ISBN": ("×今回使わない", "P3", "無", "無", "無", "N", "N", "N", "N", "N", "N", ""),
    "コードが今捨てている": ("◎使えそう", "P0", "高", "高", "高", "Y", "Y", "Y", "Y", "N", "N", "A_Keepa,B_Step2,B_Step76,CPO"),
}


def load_base():
    text = CSV_IN.read_text(encoding="utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


RAKUTEN_ITEMS = [
    # scope, field, jp, pick, pri, set, price, human, flows, note
    ("検索全体", "count", "ヒット総数", "○補助", "P1", "低", "低", "低", "A_Keepa,B_Step2", "0件判定"),
    ("検索全体", "page", "ページ番号", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("検索全体", "first", "先頭位置", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("検索全体", "last", "末尾位置", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("検索全体", "hits", "返却件数", "○補助", "P1", "低", "低", "低", "A_Keepa", "キャッシュヘッダ"),
    ("検索全体", "carrier", "PC/スマホ", "×今回使わない", "P3", "無", "無", "無", "", "入力で変わる"),
    ("検索全体", "pageCount", "総ページ", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "itemName", "商品名", "◎使えそう", "P0", "高", "中", "高", "A_Keepa,B_Step2,CPO,除外フィルタ", "袋数の主ソース"),
    ("商品", "catchcopy", "キャッチ", "○補助", "P1", "中", "低", "中", "A_Keepa,B_Step2", "名前と結合推奨（公式）"),
    ("商品", "itemCode", "商品コード", "○補助", "P1", "低", "低", "低", "A_Keepa,B_Step2", "キャッシュ安定ID"),
    ("商品", "itemPrice", "価格", "◎使えそう", "P0", "高", "高", "高", "A_Keepa,B_Step2,CPO", "表示価格"),
    ("商品", "itemCaption", "説明文", "○補助", "P1", "中", "低", "中", "A_Keepa", "袋数がタイトルに無いとき"),
    ("商品", "itemUrl", "商品URL", "○補助", "P1", "低", "低", "高", "A_Keepa,B_Step2,CPO", "人間確認"),
    ("商品", "itemPriceBaseField", "価格フィールド種別", "×今回使わない", "P3", "無", "無", "無", "", "min1/2/3のどれか"),
    ("商品", "itemPriceMax1", "価格max全体", "△後続", "P2", "低", "中", "低", "B_Step2", "選択式の幅"),
    ("商品", "itemPriceMax2", "価格max検索可", "△後続", "P2", "低", "中", "低", "B_Step2", ""),
    ("商品", "itemPriceMax3", "価格max購入可", "△後続", "P2", "低", "中", "低", "B_Step2", "選択SKUの上限"),
    ("商品", "itemPriceMin1", "価格min全体", "△後続", "P2", "低", "中", "低", "B_Step2,除外フィルタ", "500円選択の検知"),
    ("商品", "itemPriceMin2", "価格min検索可", "△後続", "P2", "低", "中", "低", "B_Step2", ""),
    ("商品", "itemPriceMin3", "価格min購入可", "○補助", "P1", "中", "中", "中", "A_Keepa,B_Step2,除外フィルタ", "選択式の下限"),
    ("商品", "affiliateUrl", "アフィリエイトURL", "×今回使わない", "P3", "無", "無", "無", "", "affiliateId時のみ"),
    ("商品", "imageFlag", "画像有無", "○補助", "P1", "低", "低", "低", "A_Keepa", "Vision対象の選別"),
    ("商品", "smallImageUrls", "画像64px×最大3", "○補助", "P1", "中", "低", "中", "A_Keepa,B_Step2", "medium優先"),
    ("商品", "mediumImageUrls", "画像128px×最大3", "◎使えそう", "P0", "高", "低", "高", "A_Keepa,B_Step2", "袋数Vision"),
    ("商品", "availability", "在庫フラグ", "○補助", "P1", "中", "低", "低", "A_Keepa,B_Step2,除外フィルタ", "0欠品1あり"),
    ("商品", "taxFlag", "税フラグ", "○補助", "P1", "低", "中", "低", "B_Step2,CPO", "0込1別。実質の補正"),
    ("商品", "postageFlag", "送料フラグ", "◎使えそう", "P0", "中", "高", "中", "A_Keepa,B_Step2,CPO,除外フィルタ", "0込1別。円は無い"),
    ("商品", "creditCardFlag", "クレカ", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "shopOfTheYearFlag", "SOTY", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "shipOverseasFlag", "海外発送", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "shipOverseasArea", "海外エリア", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "asurakuFlag", "あす楽", "×今回使わない", "P3", "無", "無", "無", "", "2024-07以降0固定"),
    ("商品", "asurakuClosingTime", "あす楽締切", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "asurakuArea", "あす楽地域", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "affiliateRate", "アフィリエイト料率", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "startTime", "セール開始", "△後続", "P2", "低", "中", "低", "CPO", "期間限定価格の信頼度"),
    ("商品", "endTime", "セール終了", "△後続", "P2", "低", "中", "低", "CPO", ""),
    ("商品", "reviewCount", "レビュー件数", "○補助", "P1", "中", "低", "中", "A_Keepa,B_Step2,B_Step76", "投票の重み"),
    ("商品", "reviewAverage", "レビュー平均", "△後続", "P2", "低", "低", "低", "B_Step2,除外フィルタ", ""),
    ("商品", "pointRate", "ポイント倍率", "◎使えそう", "P0", "中", "高", "中", "A_Keepa,B_Step2,CPO", "実質価格。24h以内終了は非表示"),
    ("商品", "pointRateStartTime", "ポイント開始", "○補助", "P1", "低", "高", "低", "A_Keepa,CPO", "キャンペーン信頼性"),
    ("商品", "pointRateEndTime", "ポイント終了", "○補助", "P1", "低", "高", "低", "A_Keepa,CPO", ""),
    ("商品", "giftFlag", "ギフト包装", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "genreId", "ジャンルID", "△後続", "P2", "低", "低", "低", "B_Step76", "Keepaカテゴリと非互換"),
    ("商品", "attributeIds", "属性ID配列", "△後続", "P2", "低", "低", "低", "B_Step76", "attributeFlag=1かつgenreId必須"),
    ("店舗", "shopName", "店舗名", "◎使えそう", "P0", "中", "低", "中", "A_Keepa,B_Step2,除外フィルタ", "ふるさと納税店の除外"),
    ("店舗", "shopCode", "店舗コード", "○補助", "P1", "低", "低", "低", "除外フィルタ", "除外リスト用"),
    ("店舗", "shopUrl", "店舗URL", "○補助", "P1", "低", "低", "低", "除外フィルタ", ""),
    ("店舗", "shopAffiliateUrl", "店舗アフィリURL", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("オプション", "GenreInformation.ancestors", "親ジャンル", "△後続", "P2", "低", "低", "低", "B_Step76", "genreInformationFlag=1"),
    ("オプション", "GenreInformation.genre", "指定ジャンル", "△後続", "P2", "低", "低", "低", "B_Step76", ""),
    ("オプション", "GenreInformation.children", "子ジャンル", "△後続", "P2", "低", "低", "低", "B_Step76", ""),
    ("オプション", "attributes[]", "属性名・ID", "△後続", "P2", "低", "低", "低", "B_Step76", "ジャンル必須。内容量属性の可能性"),
]

YAHOO_ITEMS = [
    ("検索全体", "totalResultsAvailable", "ヒット総数", "○補助", "P1", "低", "低", "低", "A_Keepa,B_Step2", "0件→queryフォールバック"),
    ("検索全体", "totalResultsReturned", "返却件数", "○補助", "P1", "低", "低", "低", "A_Keepa", ""),
    ("検索全体", "firstResultsPosition", "先頭位置", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("検索全体", "request/query", "リクエストquery", "○補助", "P1", "低", "低", "低", "A_Keepa", "キャッシュの検索条件"),
    ("商品", "hits/index", "順位", "○補助", "P1", "低", "低", "低", "A_Keepa,B_Step2", "1件目禁止の印"),
    ("商品", "hits/name", "商品名", "◎使えそう", "P0", "高", "中", "高", "A_Keepa,B_Step2,CPO,除外フィルタ", "袋数の主ソース"),
    ("商品", "hits/description", "商品説明", "○補助", "P1", "中", "低", "中", "A_Keepa", ""),
    ("商品", "hits/headLine", "キャッチ", "○補助", "P1", "中", "低", "中", "A_Keepa", "公式はheadLine（headではない）"),
    ("商品", "hits/inStock", "在庫", "○補助", "P1", "中", "低", "低", "A_Keepa,B_Step2,除外フィルタ", ""),
    ("商品", "hits/url", "商品URL", "○補助", "P1", "低", "低", "高", "A_Keepa,B_Step2,CPO", ""),
    ("商品", "hits/code", "商品コード", "○補助", "P1", "低", "低", "低", "A_Keepa,B_Step2", "seller_managed_item_id"),
    ("商品", "hits/condition", "新品/中古", "○補助", "P1", "中", "低", "低", "A_Keepa,B_Step2,除外フィルタ", "used除外"),
    ("商品", "hits/taxExcludePrice", "税抜価格", "○補助", "P1", "低", "中", "低", "CPO", "ポイント計算が税抜基準の時期あり"),
    ("商品", "hits/taxExcludePremiumPrice", "税抜プレミアム", "△後続", "P2", "低", "中", "低", "CPO", ""),
    ("商品", "hits/premiumPrice", "プレミアム価格", "△後続", "P2", "低", "中", "低", "B_Step2,CPO", "通常priceと乖離除外"),
    ("商品", "hits/premiumDiscountType", "プレミアム割引種別", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "hits/premiumDiscountRate", "プレミアム割引率", "△後続", "P2", "低", "低", "低", "CPO", ""),
    ("商品", "hits/imageId", "画像ID", "○補助", "P1", "低", "低", "低", "A_Keepa", ""),
    ("商品", "hits/image/small", "画像76px", "○補助", "P1", "中", "低", "中", "A_Keepa", ""),
    ("商品", "hits/image/medium", "画像146px", "◎使えそう", "P0", "高", "低", "高", "A_Keepa,B_Step2", "袋数Vision"),
    ("商品", "hits/exImage/*", "指定サイズ画像", "△後続", "P2", "中", "低", "中", "A_Keepa", "image_size=300/600で鮮明"),
    ("商品", "hits/review/rate", "レビュー点", "△後続", "P2", "低", "低", "低", "除外フィルタ", ""),
    ("商品", "hits/review/count", "レビュー件数", "○補助", "P1", "中", "低", "中", "A_Keepa,B_Step2,B_Step76", "売れ筋近似"),
    ("商品", "hits/review/url", "レビューURL", "×今回使わない", "P3", "無", "無", "無", "", "2026-08追加予定"),
    ("商品", "hits/affiliateRate", "アフィリ料率", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "hits/price", "価格", "◎使えそう", "P0", "高", "高", "高", "A_Keepa,B_Step2,CPO", ""),
    ("商品", "hits/priceLabel/taxable", "税込か", "○補助", "P1", "低", "中", "低", "CPO", ""),
    ("商品", "hits/priceLabel/premiumPrice", "ラベルPrem", "△後続", "P2", "低", "中", "低", "CPO", ""),
    ("商品", "hits/priceLabel/taxExcludePremiumPrice", "ラベルPrem税抜", "△後続", "P2", "低", "低", "低", "CPO", ""),
    ("商品", "hits/priceLabel/defaultPrice", "通常価格", "△後続", "P2", "低", "中", "低", "除外フィルタ", "外れ値"),
    ("商品", "hits/priceLabel/taxExcludeDefaultPrice", "通常税抜", "△後続", "P2", "低", "低", "低", "CPO", ""),
    ("商品", "hits/priceLabel/discountedPrice", "セール価格", "○補助", "P1", "低", "中", "低", "CPO", "期間限定の信頼度"),
    ("商品", "hits/priceLabel/taxExcludeDiscountedPrice", "セール税抜", "△後続", "P2", "低", "低", "低", "CPO", ""),
    ("商品", "hits/priceLabel/fixedPrice", "定価", "△後続", "P2", "低", "低", "低", "除外フィルタ", ""),
    ("商品", "hits/priceLabel/periodStart", "セール開始", "△後続", "P2", "低", "中", "低", "CPO", ""),
    ("商品", "hits/priceLabel/periodEnd", "セール終了", "△後続", "P2", "低", "中", "低", "CPO", ""),
    ("商品", "hits/point/amount", "Tポイント数", "×今回使わない", "P3", "無", "無", "無", "", "2022-04以降0固定"),
    ("商品", "hits/point/times", "Tポイント倍率", "×今回使わない", "P3", "無", "無", "無", "", "0固定"),
    ("商品", "hits/point/bonusAmount", "旧PayPayポイント", "×今回使わない", "P3", "無", "無", "無", "", "2025-02以降0固定"),
    ("商品", "hits/point/bonusTimes", "旧PayPay倍率", "×今回使わない", "P3", "無", "無", "無", "", "0固定"),
    ("商品", "hits/point/lyLimitedBonusAmount", "期間限定PayPay数", "◎使えそう", "P0", "中", "高", "中", "A_Keepa,B_Step2,CPO", "現行の実質ポイント"),
    ("商品", "hits/point/lyLimitedBonusTimes", "期間限定PayPay倍率", "◎使えそう", "P0", "中", "高", "中", "A_Keepa,B_Step2,CPO", ""),
    ("商品", "hits/point/premiumAmount", "Prem Tポイント", "×今回使わない", "P3", "無", "無", "無", "", "0固定"),
    ("商品", "hits/point/premiumTimes", "Prem T倍率", "×今回使わない", "P3", "無", "無", "無", "", "0固定"),
    ("商品", "hits/point/premiumBonusAmount", "旧Prem PayPay", "×今回使わない", "P3", "無", "無", "無", "", "0固定"),
    ("商品", "hits/point/premiumBonusTimes", "旧Prem倍率", "×今回使わない", "P3", "無", "無", "無", "", "0固定"),
    ("商品", "hits/point/lyLimitedPremiumBonusAmount", "Prem期間限定数", "△後続", "P2", "低", "中", "低", "CPO", "一般客の実質には使わない"),
    ("商品", "hits/point/lyLimitedPremiumBonusTimes", "Prem期間限定倍率", "△後続", "P2", "低", "中", "低", "CPO", ""),
    ("商品", "hits/shipping/name", "送料条件名", "○補助", "P1", "低", "中", "低", "A_Keepa,B_Step2", ""),
    ("商品", "hits/shipping/code", "送料コード", "◎使えそう", "P0", "中", "高", "中", "A_Keepa,B_Step2,CPO,除外フィルタ", "1無設定 2無料 3条件付無料。円は無い"),
    ("商品", "hits/genreCategory/id", "ジャンルID", "△後続", "P2", "低", "低", "低", "B_Step76", "SCと不一致あり"),
    ("商品", "hits/genreCategory/name", "ジャンル名", "△後続", "P2", "低", "低", "低", "B_Step76", ""),
    ("商品", "hits/genreCategory/depth", "ジャンル階層", "△後続", "P2", "低", "低", "低", "B_Step76", ""),
    ("商品", "hits/parentGenreCategories[]", "親ジャンル配列", "△後続", "P2", "低", "低", "低", "B_Step76", "パス連結"),
    ("商品", "hits/brand/id", "ブランドID", "△後続", "P2", "低", "低", "低", "B_Step76", "出品コードと別"),
    ("商品", "hits/brand/name", "ブランド名", "○補助", "P1", "低", "低", "低", "A_Keepa,B_Step76", "文字列一致"),
    ("商品", "hits/parentBrands[]", "親ブランド", "△後続", "P2", "低", "低", "低", "B_Step76", ""),
    ("商品", "hits/janCode", "JAN", "◎使えそう", "P0", "高", "中", "高", "A_Keepa,B_Step2", "Keepa EANとJOIN"),
    ("商品", "hits/payment", "支払コード", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("商品", "hits/releaseDate", "発売日", "△後続", "P2", "低", "低", "低", "A_Keepa", ""),
    ("店舗", "hits/seller/sellerId", "ストアID", "○補助", "P1", "低", "低", "低", "除外フィルタ", ""),
    ("店舗", "hits/seller/name", "ストア名", "◎使えそう", "P0", "中", "低", "中", "A_Keepa,B_Step2,除外フィルタ", ""),
    ("店舗", "hits/seller/url", "ストアURL", "○補助", "P1", "低", "低", "低", "除外フィルタ", "ロハコ注意"),
    ("店舗", "hits/seller/isBestSeller", "ベストストア", "△後続", "P2", "低", "低", "低", "B_Step2", "重みの補助"),
    ("店舗", "hits/seller/payment", "店支払（終了）", "×今回使わない", "P3", "無", "無", "無", "", "提供終了"),
    ("店舗", "hits/seller/review/rate", "店レビュー点", "△後続", "P2", "低", "低", "低", "除外フィルタ", ""),
    ("店舗", "hits/seller/review/count", "店レビュー件数", "△後続", "P2", "低", "低", "低", "B_Step2", ""),
    ("店舗", "hits/seller/imageId", "店画像ID", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("配送", "hits/delivery/area", "都道府県", "×今回使わない", "P3", "無", "無", "無", "", "delivery_*3点指定時"),
    ("配送", "hits/delivery/deadLine", "締め時間", "×今回使わない", "P3", "無", "無", "無", "", ""),
    ("配送", "hits/delivery/day", "配送日数", "×今回使わない", "P3", "無", "無", "無", "", "きょう/あす"),
]


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER
        cell.font = HFONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"


def write_rows(ws, headers, rows, col_widths):
    ws.append(headers)
    pick_col = headers.index("agent_pick") + 1 if "agent_pick" in headers else None
    pri_col = headers.index("priority") + 1 if "priority" in headers else None
    for row in rows:
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = WRAP
            cell.border = THIN
        if pick_col:
            f = fill_pick(str(ws.cell(r, pick_col).value or ""))
            if f:
                ws.cell(r, pick_col).fill = f
        if pri_col:
            pv = str(ws.cell(r, pri_col).value or "")
            if pv == "P0":
                ws.cell(r, pri_col).fill = GREEN
            elif pv == "P1":
                ws.cell(r, pri_col).fill = YELLOW
            elif pv == "P2":
                ws.cell(r, pri_col).fill = ORANGE
            elif pv == "P3":
                ws.cell(r, pri_col).fill = GRAY
    style_header(ws, len(headers))
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36


def main():
    base = load_base()
    wb = Workbook()

    # legend
    leg = wb.active
    leg.title = "凡例_ピボット"
    legend_rows = [
        ["項目", "意味", "色", "使い方"],
        ["agent_pick ◎使えそう", "A前キャッシュで残し、後続で再検索せず使う", "緑", "フィルタ=◎"],
        ["agent_pick ○補助", "残しておくと精度/除外に効く", "黄", ""],
        ["agent_pick △後続", "セット数Aには不要。カテゴリ/CPOで後から", "橙", ""],
        ["agent_pick ×今回使わない", "食品リサーチでは無視してよい", "灰", ""],
        ["agent_pick K Keepa専用", "楽天Yahooに無い。A/FBAで使う", "青", ""],
        ["priority P0", "最初の実装で必須", "緑", "ピボット行=priority"],
        ["priority P1", "本線だが2周目でも可", "黄", ""],
        ["priority P2", "メニュー8やCPO", "橙", ""],
        ["priority P3", "実装しない", "灰", ""],
        ["effect_setcount", "Keepaセット数の手間削減", "", "ピボット値=件数"],
        ["effect_price", "実質価格・競合の質", "", ""],
        ["effect_human", "人間の目視時間", "", ""],
        ["flow_tags", "カンマ区切りフロー", "", "データ→テキスト分割してピボット"],
        ["flow_A_Keepa", "メニューA（Keepa取得・セット数）", "", "Y/N列でもピボット可"],
        ["flow_B_Step2", "B統合 Step2 モール横断", "", ""],
        ["flow_B_Step76", "B Step7.6 Yahoo/楽天カテゴリ", "", ""],
        ["flow_CPO", "⑤ 楽天Yahoo価格提案", "", ""],
        ["flow_FBA", "FBA手数料・梱包", "", ""],
        ["flow_除外フィルタ", "ふるさと/選択式/中古の除外", "", ""],
        ["②完全性", "「対応表」は論理列の要約。公式の1フィールド1行は「楽天_公式全項目」「Yahoo_公式全項目」", "", "Keepa全項目はこのブックに未収録（別CSV 石原水産_Keepa項目過不足）"],
        ["楽天出典", "Ichiba Item Search 出力（2026-07-01公式。現行コードは20220601だが出力項目はほぼ同じ）", "", "https://webservice.rakuten.co.jp/documentation/ichiba-item-search"],
        ["Yahoo出典", "商品検索v3 レスポンス", "", "https://developer.yahoo.co.jp/webapi/shopping/v3/itemsearch.html"],
        ["未収録", "楽天: 入力パラメータのみの項目。属性はgenreId指定時のみ増える", "", "Yahoo: リクエスト専用（sort等）。itemLookupは別API"],
        ["コード現状", "楽天パースは name/price/url/image1/postageFlag/pointRate のみ。Yahooは name/price/url/image/shippingCode のみ", "", "◎はraw保存すれば使える"],
    ]
    for row in legend_rows:
        leg.append(row)
    style_header(leg, 4)
    for i, w in enumerate([28, 70, 10, 55], 1):
        leg.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, 7):
        leg.cell(r, 3).fill = [GREEN, YELLOW, ORANGE, GRAY, BLUE][r - 2]

    # sheet 対応表
    ws = wb.create_sheet("対応表_論理列")
    extra_h = [
        "agent_pick",
        "priority",
        "effect_setcount",
        "effect_price",
        "effect_human",
        "flow_A_Keepa",
        "flow_B_Step2",
        "flow_B_Step76",
        "flow_CPO",
        "flow_FBA",
        "flow_除外フィルタ",
        "flow_tags",
    ]
    headers = list(base[0].keys()) + extra_h
    rows = []
    for rec in base:
        key = rec.get("logical_column_ja") or ""
        t = TAGS.get(key, ("△後続", "P2", "低", "低", "低", "N", "N", "N", "N", "N", "N", ""))
        rows.append([rec.get(h, "") for h in rec.keys()] + list(t))
    write_rows(
        ws,
        headers,
        rows,
        [12, 22, 16, 36, 36, 36, 28, 40, 40, 16, 10, 14, 12, 12, 12, 12, 12, 12, 12, 36],
    )

    rh = [
        "api",
        "scope",
        "field",
        "name_ja",
        "agent_pick",
        "priority",
        "effect_setcount",
        "effect_price",
        "effect_human",
        "flow_tags",
        "note",
        "in_logical_crosswalk",
    ]
    logical_rakuten = " ".join(r.get("rakuten_ichiba") or "" for r in base)

    def in_cross(field):
        return "Y" if field.split("/")[-1].split("[")[0] in logical_rakuten or field in logical_rakuten else "N"

    wr = wb.create_sheet("楽天_公式全項目")
    rrows = []
    for it in RAKUTEN_ITEMS:
        scope, field, jp, pick, pri, s, p, h, flows, note = it
        rrows.append(
            ["楽天IchibaSearch", scope, field, jp, pick, pri, s, p, h, flows, note, in_cross(field)]
        )
    write_rows(wr, rh, rrows, [18, 14, 36, 22, 16, 10, 14, 12, 12, 40, 40, 16])

    wy = wb.create_sheet("Yahoo_公式全項目")
    yrows = []
    logical_yahoo = " ".join(r.get("yahoo_itemsearch_v3") or "" for r in base)
    for it in YAHOO_ITEMS:
        scope, field, jp, pick, pri, s, p, h, flows, note = it
        yrows.append(
            [
                "Yahoo itemSearch v3",
                scope,
                field,
                jp,
                pick,
                pri,
                s,
                p,
                h,
                flows,
                note,
                "Y" if any(x in logical_yahoo for x in field.split("/")[-2:]) else "N",
            ]
        )
    write_rows(wy, rh, yrows, [20, 14, 42, 24, 16, 10, 14, 12, 12, 40, 44, 16])

    # counts
    c = wb.create_sheet("件数サマリ")
    c.append(["シート", "行数", "◎", "○", "△", "×", "K", "公式か"])
    def count_pick(items, idx=3):
        # for tuple lists pick is index 3
        from collections import Counter
        cc = Counter()
        for it in items:
            p = it[3]
            if p.startswith("◎"):
                cc["◎"] += 1
            elif p.startswith("○"):
                cc["○"] += 1
            elif p.startswith("△"):
                cc["△"] += 1
            elif p.startswith("×"):
                cc["×"] += 1
            elif p.startswith("K"):
                cc["K"] += 1
        return cc

    cr = count_pick(RAKUTEN_ITEMS)
    cy = count_pick(YAHOO_ITEMS)
    from collections import Counter
    cl = Counter()
    for rec in base:
        p = TAGS.get(rec["logical_column_ja"], ("△",))[0]
        cl[p[0]] += 1
    c.append(["対応表_論理列", len(base), cl.get("◎", 0), cl.get("○", 0), cl.get("△", 0), cl.get("×", 0), cl.get("K", 0), "要約（全項目ではない）"])
    c.append(["楽天_公式全項目", len(RAKUTEN_ITEMS), cr["◎"], cr["○"], cr["△"], cr["×"], 0, "検索レスポンス出力はこれで全部（属性はオプション）"])
    c.append(["Yahoo_公式全項目", len(YAHOO_ITEMS), cy["◎"], cy["○"], cy["△"], cy["×"], 0, "v3 hits配下はこれで全部（0固定の旧ポイント含む）"])
    c.append(["Keepa_product辞書", "keepa_official.csv", "", "", "", "", "", "csv[]非保存。列化は◎＋K。残りは生JSON"])
    style_header(c, 8)
    for i, w in enumerate([22, 10, 8, 8, 8, 8, 8, 70], 1):
        c.column_dimensions[get_column_letter(i)].width = w

    DOCS_DIR.mkdir(parents=True, exist_ok=True)

    def dump_csv(name, headers, data_rows):
        p = DOCS_DIR / name
        with p.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(data_rows)
        print("csv", p)

    dump_csv("logical_crosswalk.csv", headers, rows)
    dump_csv("rakuten_official.csv", rh, rrows)
    dump_csv("yahoo_official.csv", rh, yrows)

    docs_xlsx = DOCS_DIR / "competitor_fields_workbook.xlsx"
    wb.save(docs_xlsx)
    print("docs xlsx", docs_xlsx)

    try:
        wb.save(OUT)
        print("saved", OUT)
    except PermissionError:
        print("tools xlsx locked, docs copy is enough")


if __name__ == "__main__":
    main()
